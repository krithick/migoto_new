# # # # # # import openpyxl
# # # # # # from openpyxl.styles import Font, PatternFill
# # # # # # from openpyxl.worksheet.filters import FilterColumn, AutoFilter
# # # # # # from openpyxl.utils import get_column_letter

# # # # # # # Create a new workbook
# # # # # # workbook = openpyxl.Workbook()

# # # # # # # Remove the default sheet
# # # # # # workbook.remove(workbook.active)

# # # # # # # Define the data for each sheet
# # # # # # sheets_data = {
# # # # # #     "Core Features": [
# # # # # #         {"Feature": "Hierarchy", "Description": "Displays and manages GameObjects.", "Backend Requirements": "API for CRUD operations on scene objects + DB storage.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Feature": "Scene", "Description": "Create, edit, and organize scenes.", "Backend Requirements": "Scene save/load API, autosave, version control, cloud sync.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Feature": "Custom Assets", "Description": "Import 2D images, 3D models, textures, videos, etc.", "Backend Requirements": "File upload APIs, asset processing, cloud storage, CDN delivery.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # # #         {"Feature": "Library", "Description": "Access to various default asset libraries.", "Backend Requirements": "Asset catalog service, tagging, search API, caching.", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Jan 2026"},
# # # # # #         {"Feature": "Step Manager", "Description": "Organize and order steps for interactions.", "Backend Requirements": "JSON-based step definitions stored in DB, reordering APIs, undo/redo support.", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Feb 2026"},
# # # # # #         {"Feature": "Add Component", "Description": "Attach custom behaviors or scripts to objects.", "Backend Requirements": "Script execution sandbox, script validation, event-driven backend integration.", "Priority": "High", "Milestone": "Phase 3", "Timeline": "Jun 2026"}
# # # # # #     ],
# # # # # #     "Sub-Features": [
# # # # # #         {"Parent Feature": "Scene", "Sub-Feature": "Move Transform", "Description": "Move a GameObject in 3D space.", "Backend Requirements": "Front-end only", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Parent Feature": "Scene", "Sub-Feature": "Rotate Transform", "Description": "Rotate a GameObject.", "Backend Requirements": "Front-end only", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Parent Feature": "Scene", "Sub-Feature": "Scale Transform", "Description": "Scale a GameObject.", "Backend Requirements": "Front-end only", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Parent Feature": "Scene", "Sub-Feature": "2D Objects", "Description": "Add 2D UI components like buttons, canvases, etc.", "Backend Requirements": "Sync with scene schema in DB", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Feb 2026"},
# # # # # #         {"Parent Feature": "Library", "Sub-Feature": "Default Library", "Description": "Basic 3D primitives.", "Backend Requirements": "Bundled locally or cached", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Jan 2026"},
# # # # # #         {"Parent Feature": "Library", "Sub-Feature": "Fire Safety Library", "Description": "Predefined fire safety 3D models.", "Backend Requirements": "Stored on CDN, API for retrieval", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "May 2026"},
# # # # # #         {"Parent Feature": "Library", "Sub-Feature": "Height Safety Library", "Description": "Predefined height safety models.", "Backend Requirements": "Stored on CDN, API for retrieval", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "May 2026"},
# # # # # #         {"Parent Feature": "Library", "Sub-Feature": "Hyundai Library", "Description": "Specialized environment models.", "Backend Requirements": "Stored on CDN, API for retrieval", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "Jun 2026"}
# # # # # #     ],
# # # # # #     "Recommended Features": [
# # # # # #         {"Feature": "User Authentication", "Description": "Sign-in, role-based access, multi-org support.", "Backend Requirements": "OAuth2, JWT, user DB, role permissions.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Feature": "Project Management", "Description": "Create, duplicate, delete projects.", "Backend Requirements": "Project schema in DB, version history, project APIs.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # # #         {"Feature": "Version Control", "Description": "Track scene versions, rollback, compare diffs.", "Backend Requirements": "Git-like model in DB or snapshotting service.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Nov 2025"},
# # # # # #         {"Feature": "Real-Time Collaboration", "Description": "Multiple users edit same scene simultaneously.", "Backend Requirements": "WebSockets/WebRTC, conflict resolution, operational transforms.", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Jan 2026"},
# # # # # #         {"Feature": "Custom Scripting", "Description": "Attach behaviors like onClick/onEnter/physics.", "Backend Requirements": "Sandbox engine, secure script executor, event bus.", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Feb 2026"},
# # # # # #         {"Feature": "Publishing System", "Description": "Export VR experiences as web/mobile builds.", "Backend Requirements": "Build service integration, CDN deployment pipeline.", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Mar 2026"},
# # # # # #         {"Feature": "Analytics Dashboard", "Description": "See project usage, interactions, errors.", "Backend Requirements": "Event tracking DB, analytics pipeline, visualization APIs.", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Jun 2026"},
# # # # # #         {"Feature": "Asset Optimization", "Description": "Automatic compression, LOD generation, light baking.", "Backend Requirements": "Background workers for asset processing.", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "May 2026"},
# # # # # #         {"Feature": "Preview & Testing", "Description": "Device previews (Oculus, WebXR, ARKit/ARCore).", "Backend Requirements": "Streaming service or APK build APIs.", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Jul 2026"},
# # # # # #         {"Feature": "Localization Support", "Description": "Multilingual asset metadata and UI labels.", "Backend Requirements": "Translation DB, key-value management.", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "Aug 2026"},
# # # # # #         {"Feature": "AI Assistance (Optional)", "Description": "Auto-generate scripts, recommend assets.", "Backend Requirements": "Integrate OpenAI/LLM APIs.", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "Aug 2026"}
# # # # # #     ],
# # # # # #     "Backend Architecture": [
# # # # # #         {"Component": "API Gateway", "Description": "Routes requests to relevant services.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Component": "Auth Service", "Description": "Manages authentication, roles, permissions.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Component": "Asset Service", "Description": "Handles upload, optimization, and serving of assets.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # # #         {"Component": "Scene Service", "Description": "CRUD operations, autosave, and version control for scenes.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # # #         {"Component": "Project Service", "Description": "Manages project metadata and collaboration.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Nov 2025"},
# # # # # #         {"Component": "Publishing Service", "Description": "Handles builds and deployment of VR projects.", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Mar 2026"},
# # # # # #         {"Component": "Analytics Service", "Description": "Collects and visualizes usage metrics.", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Jun 2026"},
# # # # # #         {"Component": "Database", "Description": "MongoDB for scene graphs, Postgres for metadata, Redis for caching.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # # #         {"Component": "Cloud Storage", "Description": "S3/GCS/Azure Blob for assets.", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # # #         {"Component": "WebSocket Layer", "Description": "Supports real-time collaboration.", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Jan 2026"},
# # # # # #         {"Component": "CI/CD Pipeline", "Description": "Automated deployment and publishing of builds.", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Mar 2026"}
# # # # # #     ]
# # # # # # }

# # # # # # # Priority order for sorting
# # # # # # priority_order = {"High": 1, "Medium": 2, "Low": 3}

# # # # # # # Create each sheet
# # # # # # for sheet_name, data in sheets_data.items():
# # # # # #     # Create a new sheet
# # # # # #     ws = workbook.create_sheet(title=sheet_name)
    
# # # # # #     # Define headers based on the first dictionary's keys
# # # # # #     headers = list(data[0].keys())
# # # # # #     for col_idx, header in enumerate(headers, start=1):
# # # # # #         cell = ws.cell(row=1, column=col_idx)
# # # # # #         cell.value = header
# # # # # #         cell.font = Font(bold=True)  # Bold headers
# # # # # #         cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue fill
    
# # # # # #     # Sort data by Priority
# # # # # #     sorted_data = sorted(data, key=lambda x: priority_order.get(x["Priority"], 3))
    
# # # # # #     # Write data to the sheet
# # # # # #     for row_idx, row_data in enumerate(sorted_data, start=2):
# # # # # #         for col_idx, key in enumerate(headers, start=1):
# # # # # #             cell = ws.cell(row=row_idx, column=col_idx)
# # # # # #             cell.value = row_data[key]
# # # # # #             cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")  # Light gray fill for data rows
    
# # # # # #     # Apply filters to the header row
# # # # # #     ws.auto_filter.ref = ws.dimensions
    
# # # # # #     # Auto-adjust column widths
# # # # # #     for col_idx in range(1, len(headers) + 1):
# # # # # #         max_length = 0
# # # # # #         column_letter = get_column_letter(col_idx)
# # # # # #         for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
# # # # # #             for cell in row:
# # # # # #                 try:
# # # # # #                     if len(str(cell.value)) > max_length:
# # # # # #                         max_length = len(str(cell.value))
# # # # # #                 except:
# # # # # #                     pass
# # # # # #         adjusted_width = (max_length + 2) * 1.2  # Add padding and adjust
# # # # # #         ws.column_dimensions[column_letter].width = adjusted_width

# # # # # # # Save the workbook
# # # # # # workbook.save("VR_Authoring_Tool_Roadmap.xlsx")
# # # # # # print("Excel file 'VR_Authoring_Tool_Roadmap.xlsx' has been created successfully.")
# # # # # import openpyxl
# # # # # from openpyxl.styles import Font, PatternFill
# # # # # from openpyxl.worksheet.filters import FilterColumn, AutoFilter
# # # # # from openpyxl.utils import get_column_letter

# # # # # # Create a new workbook
# # # # # workbook = openpyxl.Workbook()

# # # # # # Remove the default sheet
# # # # # workbook.remove(workbook.active)

# # # # # # Priority order for sorting
# # # # # priority_order = {"High": 1, "Medium": 2, "Low": 3}

# # # # # # Data for Features & Milestones sheet (merged from provided document, with consolidations, additions, and new columns)
# # # # # # Consolidated: Merged colliders into "Colliders", rendering into "Rendering", line into "Line Renderer"
# # # # # # Removed redundancies: e.g., "Render Components" and "Mesh Renderer" -> "Rendering"
# # # # # # Added new features: User Authentication, Real-Time Collaboration, Version Control, Analytics Dashboard, Localization Support
# # # # # # Assigned priorities, timelines (based on current date Aug 26, 2025; completed = Aug 2025, Milestone 2 = Sept-Dec 2025, Milestone 3 = Jan 2026+)
# # # # # # Backend Requirements: Inferred/reasoned based on feature needs
# # # # # # Dependencies and Notes: Added logical ones
# # # # # features_milestones_data = [
# # # # #     {"Sno": 1, "Feature": "Hierarchy", "Description": "To Display the GameObjects", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "API for CRUD operations on scene objects + DB storage", "Dependencies": "None", "Notes": "Core feature for scene management"},
# # # # #     {"Sno": 2, "Feature": "Scene", "Description": "Logic Handling Environment", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Scene save/load API, autosave, version control, cloud sync", "Dependencies": "Hierarchy", "Notes": ""},
# # # # #     {"Sno": 3, "Feature": "Move Transform", "Description": "To move an GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Front-end only", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 4, "Feature": "Rotate Transform", "Description": "To Rotate an GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Front-end only", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 5, "Feature": "Scale Transform", "Description": "To Scale the GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Front-end only", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 6, "Feature": "Custom Assets", "Description": "To import the custom 2d images, 3d Models and Audio", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "File upload APIs, asset processing, cloud storage, CDN delivery", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 7, "Feature": "Library", "Description": "To access the default Assets", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Asset catalog service, tagging, search API, caching", "Dependencies": "None", "Notes": ""},
# # # # #     {"Sno": 8, "Feature": "2D Objects", "Description": "Canvas, Line Renderer, Button and So on", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Sync with scene schema in DB", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 9, "Feature": "Fire Safety Library", "Description": "Default Fire Safety 3d Models", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Stored on CDN, API for retrieval", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 10, "Feature": "Default Library", "Description": "Common 3d Models Like Cube /Capsule", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Bundled locally or cached", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 11, "Feature": "Height Safety Library", "Description": "Default Height Safety 3d Models", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Stored on CDN, API for retrieval", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 12, "Feature": "Hyundai Library", "Description": "Default Hyundai Environment 3d Models", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Stored on CDN, API for retrieval", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 13, "Feature": "Step Manager", "Description": "To order the logics with Drag and change feature", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "JSON-based step definitions stored in DB, reordering APIs, undo/redo support", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 14, "Feature": "Add Component", "Description": "To Add components if needed", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Script execution sandbox, script validation, event-driven backend integration", "Dependencies": "Hierarchy", "Notes": ""},
# # # # #     {"Sno": 15, "Feature": "Physics Components", "Description": "Controlling the physics of the Object Like RigidBody and Gravity", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Physics engine integration, simulation APIs", "Dependencies": "Add Component", "Notes": ""},
# # # # #     {"Sno": 16, "Feature": "Colliders", "Description": "Colliders for GameObjects (Mesh, Box, Sphere, Capsule)", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Collision detection APIs", "Dependencies": "Physics Components", "Notes": "Consolidated from individual colliders"},
# # # # #     {"Sno": 17, "Feature": "Rendering", "Description": "To render GameObjects and Meshes", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Rendering pipeline, shader support", "Dependencies": "Scene", "Notes": "Consolidated from Render Components and Mesh Renderer"},
# # # # #     {"Sno": 18, "Feature": "Interactables", "Description": "User interactions with GameObjects (click, grab, trigger)", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Event-driven interaction APIs", "Dependencies": "Add Component", "Notes": "Clarified description"},
# # # # #     {"Sno": 19, "Feature": "Audio", "Description": "Custom Audio importation", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Audio file processing, playback APIs", "Dependencies": "Custom Assets", "Notes": ""},
# # # # #     {"Sno": 20, "Feature": "HotSpot", "Description": "To show the text and active when the user hits the hotspot", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Trigger event APIs", "Dependencies": "Interactables", "Notes": ""},
# # # # #     {"Sno": 21, "Feature": "Attach to player", "Description": "To follow the player", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Attachment logic in DB", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 22, "Feature": "Climb", "Description": "Climbing feature", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Physics-based climb simulation", "Dependencies": "Physics Components", "Notes": ""},
# # # # #     {"Sno": 23, "Feature": "Line Renderer", "Description": "Line renderer", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Rendering APIs for lines", "Dependencies": "Rendering", "Notes": "Consolidated from Line and Line Renderer"},
# # # # #     {"Sno": 24, "Feature": "Lever", "Description": "Lever Feature", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Interaction APIs for levers", "Dependencies": "Interactables", "Notes": ""},
# # # # #     {"Sno": 25, "Feature": "Knob", "Description": "Knob rotation feature", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Rotation interaction APIs", "Dependencies": "Interactables", "Notes": ""},
# # # # #     {"Sno": 26, "Feature": "Grid", "Description": "To enable and Disable the Grid with Snapping feature", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Front-end only", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 27, "Feature": "Perspective", "Description": "Perspective view of the Scene", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Camera API", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 28, "Feature": "Orthographic", "Description": "Orthographic view of the Scene", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Camera API", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 29, "Feature": "Child Gameobject", "Description": "Custom Child object of the parent", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Hierarchy nesting in DB", "Dependencies": "Hierarchy", "Notes": ""},
# # # # #     {"Sno": 30, "Feature": "Offline load and Save", "Description": "Load and save mode in offline", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Local storage APIs", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 31, "Feature": "Disable and Enable GO", "Description": "Hierarchy object hide and disable", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "State management APIs", "Dependencies": "Hierarchy", "Notes": ""},
# # # # #     {"Sno": 32, "Feature": "Drag Drop", "Description": "Drag and drop from project into scene and Hierarchy", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Front-end only", "Dependencies": "Hierarchy, Scene", "Notes": ""},
# # # # #     {"Sno": 33, "Feature": "Search", "Description": "Library Search and Project Search", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Search API", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 34, "Feature": "Thumbnails", "Description": "Project Thumbnails & Library Thumbnails – Display respective environment thumbnails in the project list", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Thumbnail generation APIs", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 35, "Feature": "Copy Paste the Component Values", "Description": "Transform Copy & Paste functionality in inspector - Copy position, rotation, and scale values from one object and paste to another.", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Clipboard APIs", "Dependencies": "Add Component", "Notes": ""},
# # # # #     {"Sno": 36, "Feature": "Camera Rotation", "Description": "Allows users to rotate the camera using the right mouse button", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Front-end only", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 37, "Feature": "Performance Tracking Bar", "Description": "Bar for displaying amount of meshes are allowed to use and amount of meshes already used for transmitting through Json.", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Performance metrics APIs", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 38, "Feature": "Loading Screen", "Description": "Loading screen for login, project fetch, project load, asset importing and saving project", "Team": "CFT", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Progress tracking APIs", "Dependencies": "None", "Notes": ""},
# # # # #     {"Sno": 39, "Feature": "Options Like Hide and Lock", "Description": "Option for Hiding Inspector, Hierarchy and Asset Panels", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "UI state management", "Dependencies": "Hierarchy", "Notes": "Fixed typo in description"},
# # # # #     {"Sno": 40, "Feature": "Prevent Multiple Players", "Description": "Only one player can be added per scene", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Validation APIs", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 41, "Feature": "Sensitivity", "Description": "Sensitivity for mouse controls", "Team": "XR", "Status": "Completed", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Input configuration", "Dependencies": "Camera Rotation", "Notes": "Fixed typo"},
# # # # #     {"Sno": 42, "Feature": "VR hand Animation", "Description": "We can give normal grip, pinch hand animation for grip, and trigger button.", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Animation APIs", "Dependencies": "Interactables", "Notes": ""},
# # # # #     {"Sno": 43, "Feature": "Save", "Description": "Explicit Save Button in UI panel", "Team": "XR", "Status": "Completed", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Save API", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 44, "Feature": "Reset", "Description": "Transform position and rotation reset to zero", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025", "Backend Requirements": "Reset APIs", "Dependencies": "Move Transform", "Notes": ""},
# # # # #     {"Sno": 45, "Feature": "Cross Platform", "Description": "PC- desktop application", "Team": "XR", "Status": "", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Sept 2025", "Backend Requirements": "Build service integration for multi-platform", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 46, "Feature": "Custom scripts", "Description": "eg: move, rotate, scale, follow etc", "Team": "XR", "Status": "", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Oct 2025", "Backend Requirements": "Sandbox engine, secure script executor, event bus", "Dependencies": "Add Component", "Notes": ""},
# # # # #     {"Sno": 47, "Feature": "Multi tenancy", "Description": "Creating and managing company profiles and customer profiles", "Team": "WEB DEV", "Status": "", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Nov 2025", "Backend Requirements": "OAuth2, JWT, user DB, role permissions", "Dependencies": "User Authentication", "Notes": ""},
# # # # #     {"Sno": 48, "Feature": "RBAC", "Description": "Admin, editor(creator) and Viewer", "Team": "WEB DEV", "Status": "", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Nov 2025", "Backend Requirements": "Role-based access control APIs", "Dependencies": "Multi tenancy", "Notes": ""},
# # # # #     {"Sno": 49, "Feature": "Asset Library Creation", "Description": "Import and use our library of assets in modules", "Team": "CFT", "Status": "", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Dec 2025", "Backend Requirements": "Asset upload and catalog APIs", "Dependencies": "Library", "Notes": ""},
# # # # #     {"Sno": 50, "Feature": "Custom Assets Without restrictions", "Description": ".FBX, GLB, GLTF, any file size", "Team": "CFT", "Status": "", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Dec 2025", "Backend Requirements": "Advanced file processing, no size limits", "Dependencies": "Custom Assets", "Notes": ""},
# # # # #     {"Sno": 51, "Feature": "Viewer App Enhancement", "Description": "Viewer Invitation logic, viewer can provide comments and feedbacks.", "Team": "WEB & XR", "Status": "", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Dec 2025", "Backend Requirements": "Feedback APIs, invitation system", "Dependencies": "RBAC", "Notes": ""},
# # # # #     {"Sno": 52, "Feature": "Animation", "Description": "Keyframe based animation", "Team": "", "Status": "", "Priority": "High", "Milestone": "Phase 3", "Timeline": "Jan 2026", "Backend Requirements": "Keyframe processing, animation state machine", "Dependencies": "Custom scripts", "Notes": "Requires third-party integration"},
# # # # #     {"Sno": 53, "Feature": "Assess mode", "Description": "Build a module with analytics and report generation", "Team": "", "Status": "", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Feb 2026", "Backend Requirements": "Analytics pipeline, report APIs", "Dependencies": "Analytics Dashboard", "Notes": ""},
# # # # #     {"Sno": 54, "Feature": "Cross Platform", "Description": "Web", "Team": "", "Status": "", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Mar 2026", "Backend Requirements": "WebXR support", "Dependencies": "Cross Platform (PC)", "Notes": ""},
# # # # #     {"Sno": 55, "Feature": "Asset store", "Description": "Purchase and use our library of assets in modules", "Team": "", "Status": "", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "Apr 2026", "Backend Requirements": "E-commerce integration, asset delivery APIs", "Dependencies": "Asset Library Creation", "Notes": ""},
# # # # #     {"Sno": 56, "Feature": "Templates", "Description": "Industry and Use case wise template", "Team": "", "Status": "", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "May 2026", "Backend Requirements": "Template storage and load APIs", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 57, "Feature": "UGC integration", "Description": "Third parties should be able to add assets into the library.", "Team": "", "Status": "", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "Jun 2026", "Backend Requirements": "User-generated content APIs, moderation", "Dependencies": "Asset store", "Notes": ""},
# # # # #     # New added features
# # # # #     {"Sno": 58, "Feature": "User Authentication", "Description": "Sign-in, role-based access, multi-org support", "Team": "WEB DEV", "Status": "Planned", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Sept 2025", "Backend Requirements": "OAuth2, JWT, user DB, role permissions", "Dependencies": "None", "Notes": "Essential for security"},
# # # # #     {"Sno": 59, "Feature": "Real-Time Collaboration", "Description": "Multiple users edit same scene simultaneously", "Team": "WEB & XR", "Status": "Planned", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Oct 2025", "Backend Requirements": "WebSockets/WebRTC, conflict resolution, operational transforms", "Dependencies": "Scene, User Authentication", "Notes": ""},
# # # # #     {"Sno": 60, "Feature": "Version Control", "Description": "Track scene versions, rollback, compare diffs", "Team": "XR", "Status": "Planned", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Nov 2025", "Backend Requirements": "Git-like model in DB or snapshotting service", "Dependencies": "Scene", "Notes": ""},
# # # # #     {"Sno": 61, "Feature": "Analytics Dashboard", "Description": "See project usage, interactions, errors", "Team": "WEB DEV", "Status": "Planned", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Feb 2026", "Backend Requirements": "Event tracking DB, analytics pipeline, visualization APIs", "Dependencies": "Assess mode", "Notes": ""},
# # # # #     {"Sno": 62, "Feature": "Localization Support", "Description": "Multilingual asset metadata and UI labels", "Team": "CFT", "Status": "Planned", "Priority": "Low", "Milestone": "Phase 3", "Timeline": "Jun 2026", "Backend Requirements": "Translation DB, key-value management", "Dependencies": "UI features", "Notes": "For global users"}
# # # # # ]

# # # # # # Data for Sub-Features sheet (broken down from complex features)
# # # # # sub_features_data = [
# # # # #     {"Parent Feature": "Scene", "Sub-Feature": "Move Transform", "Description": "Move a GameObject in 3D space", "Backend Requirements": "Front-end only", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Scene", "Sub-Feature": "Rotate Transform", "Description": "Rotate a GameObject", "Backend Requirements": "Front-end only", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Scene", "Sub-Feature": "Scale Transform", "Description": "Scale a GameObject", "Backend Requirements": "Front-end only", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Scene", "Sub-Feature": "2D Objects", "Description": "Add 2D UI components like buttons, canvases, etc.", "Backend Requirements": "Sync with scene schema in DB", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Library", "Sub-Feature": "Default Library", "Description": "Basic 3D primitives", "Backend Requirements": "Bundled locally or cached", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Library", "Sub-Feature": "Fire Safety Library", "Description": "Predefined fire safety 3D models", "Backend Requirements": "Stored on CDN, API for retrieval", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Library", "Sub-Feature": "Height Safety Library", "Description": "Predefined height safety models", "Backend Requirements": "Stored on CDN, API for retrieval", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Library", "Sub-Feature": "Hyundai Library", "Description": "Specialized environment models", "Backend Requirements": "Stored on CDN, API for retrieval", "Priority": "Low", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Colliders", "Sub-Feature": "Mesh Collider", "Description": "Collider for Mesh", "Backend Requirements": "Collision APIs", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Colliders", "Sub-Feature": "Box Collider", "Description": "Box like Collider for Gameobjects", "Backend Requirements": "Collision APIs", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Colliders", "Sub-Feature": "Sphere Collider", "Description": "Sphere like collider for Gameobjects", "Backend Requirements": "Collision APIs", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     {"Parent Feature": "Colliders", "Sub-Feature": "Capsule Collider", "Description": "Capsule Like collider for Gameobjects", "Backend Requirements": "Collision APIs", "Priority": "Medium", "Milestone": "Phase 1", "Timeline": "Aug 2025"},
# # # # #     # Add more sub-features as needed
# # # # # ]

# # # # # # Data for Backend Architecture sheet (from previous recommendations)
# # # # # backend_architecture_data = [
# # # # #     {"Component": "API Gateway", "Description": "Routes requests to relevant services", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # #     {"Component": "Auth Service", "Description": "Manages authentication, roles, permissions", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # #     {"Component": "Asset Service", "Description": "Handles upload, optimization, and serving of assets", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # #     {"Component": "Scene Service", "Description": "CRUD operations, autosave, and version control for scenes", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # #     {"Component": "Project Service", "Description": "Manages project metadata and collaboration", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Nov 2025"},
# # # # #     {"Component": "Publishing Service", "Description": "Handles builds and deployment of VR projects", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Mar 2026"},
# # # # #     {"Component": "Analytics Service", "Description": "Collects and visualizes usage metrics", "Priority": "Medium", "Milestone": "Phase 3", "Timeline": "Jun 2026"},
# # # # #     {"Component": "Database", "Description": "MongoDB for scene graphs, Postgres for metadata, Redis for caching", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Sept 2025"},
# # # # #     {"Component": "Cloud Storage", "Description": "S3/GCS/Azure Blob for assets", "Priority": "High", "Milestone": "Phase 1", "Timeline": "Oct 2025"},
# # # # #     {"Component": "WebSocket Layer", "Description": "Supports real-time collaboration", "Priority": "High", "Milestone": "Phase 2", "Timeline": "Jan 2026"},
# # # # #     {"Component": "CI/CD Pipeline", "Description": "Automated deployment and publishing of builds", "Priority": "Medium", "Milestone": "Phase 2", "Timeline": "Mar 2026"}
# # # # # ]

# # # # # # Define sheets and their data
# # # # # sheets_data = {
# # # # #     "Features & Milestones": features_milestones_data,
# # # # #     "Sub-Features": sub_features_data,
# # # # #     "Backend Architecture": backend_architecture_data
# # # # # }

# # # # # # Create each sheet
# # # # # for sheet_name, data in sheets_data.items():
# # # # #     # Create a new sheet
# # # # #     ws = workbook.create_sheet(title=sheet_name)
    
# # # # #     # Define headers based on the first dictionary's keys
# # # # #     if data:  # Check if data is not empty
# # # # #         headers = list(data[0].keys())
# # # # #         for col_idx, header in enumerate(headers, start=1):
# # # # #             cell = ws.cell(row=1, column=col_idx)
# # # # #             cell.value = header
# # # # #             cell.font = Font(bold=True)  # Bold headers
# # # # #             cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue fill
    
# # # # #         # Sort data by Priority
# # # # #         sorted_data = sorted(data, key=lambda x: priority_order.get(x.get("Priority", "Low"), 3))
        
# # # # #         # Write data to the sheet
# # # # #         for row_idx, row_data in enumerate(sorted_data, start=2):
# # # # #             for col_idx, key in enumerate(headers, start=1):
# # # # #                 cell = ws.cell(row=row_idx, column=col_idx)
# # # # #                 cell.value = row_data[key]
# # # # #                 cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")  # Light gray fill for data rows
    
# # # # #         # Apply filters to the header row
# # # # #         ws.auto_filter.ref = ws.dimensions
    
# # # # #         # Auto-adjust column widths
# # # # #         for col_idx in range(1, len(headers) + 1):
# # # # #             max_length = 0
# # # # #             column_letter = get_column_letter(col_idx)
# # # # #             for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
# # # # #                 for cell in row:
# # # # #                     try:
# # # # #                         if len(str(cell.value)) > max_length:
# # # # #                             max_length = len(str(cell.value))
# # # # #                     except:
# # # # #                         pass
# # # # #             adjusted_width = (max_length + 2) * 1.2  # Add padding and adjust
# # # # #             ws.column_dimensions[column_letter].width = adjusted_width

# # # # # # Save the workbook
# # # # # workbook.save("Updated_VR_Authoring_Tool_Roadmapppp.xlsx")
# # # # # print("Excel file 'Updated_VR_Authoring_Tool_Roadmap.xlsx' has been created successfully.")
# # # # import openpyxl
# # # # from openpyxl.styles import Font, PatternFill
# # # # from openpyxl.worksheet.filters import AutoFilter
# # # # from openpyxl.utils import get_column_letter

# # # # # Create a new workbook
# # # # workbook = openpyxl.Workbook()
# # # # workbook.remove(workbook.active)

# # # # # Define priority order for sorting
# # # # priority_order = {"High": 1, "Medium": 2, "Low": 3}

# # # # # Data for Unity Features
# # # # unity_features = [
# # # #     {"Feature": "Immersive Scene Editor", "Description": "Real-time VR editing mode for creating/editing scenes in VR.", "Unity Implementation": "XR Interaction Toolkit for hand tracking; VR camera rigs.", "Cross-Platform Notes": "Adaptive for Quest/SteamVR/PC; desktop fallback.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Core for intuitive creator experience."},
# # # #     {"Feature": "Advanced Asset Import", "Description": "Support GLTF/GLB/FBX/OBJ with auto-optimization (LOD, compression).", "Unity Implementation": "AssetPostprocessor; Unity Asset Store API.", "Cross-Platform Notes": "WebGL-compatible formats.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Ensure large file support."},
# # # #     {"Feature": "Interactable System", "Description": "Pre-built interactions: grab, throw, UI pointers, haptics.", "Unity Implementation": "XRGrabInteractable, XRSimpleInteractable.", "Cross-Platform Notes": "Input abstraction for controllers/keyboard.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Scene Service", "Notes": "Test haptics on Quest/Vive."},
# # # #     {"Feature": "Animation & Keyframing", "Description": "Timeline-based animation for objects/characters.", "Unity Implementation": "Unity Timeline & Animator; Mixamo integration.", "Cross-Platform Notes": "WebXR-compatible exports.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Asset Service", "Notes": "Requires rigging support."},
# # # #     {"Feature": "Physics & Simulation", "Description": "Rigidbody, soft-body, fluid sims; raycasting.", "Unity Implementation": "Unity Physics (DOTS); NVIDIA PhysX.", "Cross-Platform Notes": "Optimize for mobile VR.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Performance critical for Quest."},
# # # #     {"Feature": "Visual Scripting", "Description": "Node-based scripting for no-code behaviors.", "Unity Implementation": "Unity Visual Scripting; custom VR nodes.", "Cross-Platform Notes": "Cross-platform node execution.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Accessible for non-coders."},
# # # #     {"Feature": "Real-Time Lighting", "Description": "Global illumination, HDRP/URP, light baking.", "Unity Implementation": "Universal Render Pipeline; adaptive quality.", "Cross-Platform Notes": "Dynamic scaling for WebXR.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Optimize for 90 FPS."},
# # # #     {"Feature": "Hand & Gesture Tracking", "Description": "Hand poses, pinch/grab, finger tracking.", "Unity Implementation": "Oculus Integration; OpenXR.", "Cross-Platform Notes": "Controller emulation fallback.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Interactable System", "Notes": "Test on Quest 3."},
# # # #     {"Feature": "Multi-User Collaboration", "Description": "Real-time co-editing with avatars, voice chat.", "Unity Implementation": "Photon Networking; Unity Netcode.", "Cross-Platform Notes": "VR/PC via WebSockets.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Collaboration Service", "Notes": "Conflict resolution critical."},
# # # #     {"Feature": "AI-Assisted Creation", "Description": "Auto-generate assets/layouts; smart suggestions.", "Unity Implementation": "ML-Agents; external AI API plugins.", "Cross-Platform Notes": "Offline for desktop; cloud for heavy compute.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Explore Stable Diffusion."},
# # # #     {"Feature": "Preview & Testing Suite", "Description": "In-editor device previews; performance profiling.", "Unity Implementation": "Unity Device Simulator; build pipelines.", "Cross-Platform Notes": "Test input variances.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Include WebXR emulator."},
# # # #     {"Feature": "Accessibility Features", "Description": "Color-blind modes, subtitles, scalable UI.", "Unity Implementation": "Unity UI Toolkit; accessibility plugins.", "Cross-Platform Notes": "WCAG compliance.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "None", "Notes": "Voice command support."},
# # # #     {"Feature": "Custom Scripting (C#)", "Description": "IDE integration for advanced scripting.", "Unity Implementation": "Visual Studio; VR script templates.", "Cross-Platform Notes": "Compile-time checks.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Visual Scripting", "Notes": "Sandbox for security."},
# # # #     {"Feature": "Audio Spatialization", "Description": "3D audio with occlusion/reverb; MIDI support.", "Unity Implementation": "Unity Audio Spatializer; FMOD.", "Cross-Platform Notes": "WebAudio for browsers.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Asset Service", "Notes": "Test on spatial audio headsets."},
# # # #     {"Feature": "Particle & Effects System", "Description": "VFX for fire/smoke; custom shaders.", "Unity Implementation": "VFX Graph; Shader Graph.", "Cross-Platform Notes": "Optimize for mobile VR.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Real-Time Lighting", "Notes": "Performance critical."}
# # # # ]

# # # # # Data for Backend Architecture
# # # # backend_architecture = [
# # # #     {"Component": "API Gateway", "Description": "Routes requests; handles auth/load balancing.", "Technologies": "AWS API Gateway; OAuth2/JWT.", "Integration with Unity": "Unity WWW/UnityWebRequest.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Ensure low latency."},
# # # #     {"Component": "Auth Service", "Description": "User sign-in, RBAC, multi-tenancy.", "Technologies": "Firebase Auth; Keycloak; SSO.", "Integration with Unity": "Unity Authentication SDK.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "GDPR compliance."},
# # # #     {"Component": "Asset Service", "Description": "Upload/process/serve assets; CDN delivery.", "Technologies": "S3/Azure Blob; FFmpeg.", "Integration with Unity": "Async uploads from Unity.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "Handle large files."},
# # # #     {"Component": "Scene Service", "Description": "CRUD for scenes; version control, autosave.", "Technologies": "MongoDB; Git-like diffs.", "Integration with Unity": "Serialize/deserialize scenes.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "Support large scene graphs."},
# # # #     {"Component": "Collaboration Service", "Description": "Real-time sync for multi-user editing.", "Technologies": "WebSockets (Socket.io); OT.", "Integration with Unity": "Unity Netcode integration.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Conflict resolution critical."},
# # # #     {"Component": "Analytics Service", "Description": "Track usage, errors, engagement; reports.", "Technologies": "Elasticsearch/Kibana.", "Integration with Unity": "Unity Analytics SDK.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "API Gateway", "Notes": "Real-time dashboards."},
# # # #     {"Component": "Publishing Service", "Description": "Build/deploy VR apps (APK, WebGL, EXE).", "Technologies": "Jenkins; Unity Cloud Build.", "Integration with Unity": "Automated builds from editor.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "Support WebGL exports."},
# # # #     {"Component": "Database Layer", "Description": "Store users/projects/assets; caching.", "Technologies": "Postgres; MongoDB; Redis.", "Integration with Unity": "ORM for queries.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "High availability."},
# # # #     {"Component": "AI Integration", "Description": "Backend for AI asset generation.", "Technologies": "Azure OpenAI; Hugging Face.", "Integration with Unity": "API calls from scripts.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Optional for pro users."},
# # # #     {"Component": "Security & Compliance", "Description": "Encryption, GDPR, vulnerability scans.", "Technologies": "HTTPS; OWASP practices.", "Integration with Unity": "Unity security plugins.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Regular audits."},
# # # #     {"Component": "Scalability Layer", "Description": "Auto-scaling; load testing.", "Technologies": "Kubernetes; Prometheus.", "Integration with Unity": "Handle session spikes.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Test for 1,000 users."},
# # # #     {"Component": "Notification Service", "Description": "Push alerts for collaboration/builds.", "Technologies": "Firebase Cloud Messaging.", "Integration with Unity": "Unity notifications API.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "API Gateway", "Notes": "User-configurable."}
# # # # ]

# # # # # Data for Cross-Platform Requirements
# # # # cross_platform = [
# # # #     {"Requirement": "Build Targets", "Description": "Support PC, VR (Quest/Vive), WebXR, Mobile AR.", "Implementation Details": "Unity Build Settings; platform-specific modules.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Test on Quest 2/3, Vive Pro."},
# # # #     {"Requirement": "Input Abstraction", "Description": "Unified input for controllers, hands, keyboard.", "Implementation Details": "OpenXR for VR; Unity Input System.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Fallback for non-VR devices."},
# # # #     {"Requirement": "Performance Optimization", "Description": "Dynamic LOD, occlusion culling.", "Implementation Details": "Unity URP; LODGroup components.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Real-Time Lighting", "Notes": "Target 90 FPS on Quest."},
# # # #     {"Requirement": "Asset Compatibility", "Description": "GLTF exports for web; poly limits for mobile.", "Implementation Details": "GLTFast plugin; asset preprocessing.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Max 50k polys for mobile."},
# # # #     {"Requirement": "Multi-Device Testing", "Description": "Emulators and beta testing for diverse hardware.", "Implementation Details": "Unity Device Simulator; external beta platforms.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Preview & Testing Suite", "Notes": "Include iOS ARKit, Android ARCore."}
# # # # ]

# # # # # Data for Product Launch Plan
# # # # launch_plan = [
# # # #     {"Task": "Define Target Audience", "Description": "Identify users (educators, designers, enterprises).", "Details": "Market research; user personas.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Nov 2025", "Dependencies": "None", "Notes": "Focus on training/simulations."},
# # # #     {"Task": "Beta Testing", "Description": "Test with 100+ users; gather feedback.", "Details": "Surveys via Typeform; beta on Steam/Oculus.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Preview & Testing Suite", "Notes": "Iterate based on feedback."},
# # # #     {"Task": "Monetization Strategy", "Description": "Freemium model; paid pro features; subscriptions.", "Details": "Unity Asset Store integration; tiered pricing.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Auth Service", "Notes": "Explore in-app purchases."},
# # # #     {"Task": "Compliance & Accessibility", "Description": "GDPR, WCAG, HIPAA compliance; accessibility audits.", "Details": "Legal consultation; automated accessibility tools.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Security & Compliance", "Notes": "Critical for enterprise."},
# # # #     {"Task": "Marketing Campaign", "Description": "Launch on Steam, Oculus Store, web; social media.", "Details": "X posts; influencer demos.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Target VR communities."},
# # # #     {"Task": "Post-Launch Support", "Description": "Updates, bug fixes, community forums.", "Details": "Discord/Reddit for support; patch pipeline.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "Monitor churn rates."},
# # # #     {"Task": "Scalability Testing", "Description": "Load test for 1,000 concurrent users.", "Details": "AWS Stress Testing; Prometheus monitoring.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scalability Layer", "Notes": "Ensure global access."}
# # # # ]

# # # # # Data for Completed Features (from original document, Milestone 1)
# # # # completed_features = [
# # # #     {"Sno": 1, "Feature": "Hierarchy", "Description": "Displays GameObjects", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Core feature completed."},
# # # #     {"Sno": 2, "Feature": "Scene", "Description": "Logic Handling Environment", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Core feature completed."},
# # # #     {"Sno": 3, "Feature": "Move Transform", "Description": "Move a GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic transform completed."},
# # # #     {"Sno": 4, "Feature": "Rotate Transform", "Description": "Rotate a GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic transform completed."},
# # # #     {"Sno": 5, "Feature": "Scale Transform", "Description": "Scale a GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic transform completed."},
# # # #     {"Sno": 6, "Feature": "Custom Assets", "Description": "Import 2D images, 3D models, audio", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Asset Service", "Notes": "Basic import completed."},
# # # #     {"Sno": 7, "Feature": "Library", "Description": "Access default assets", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Asset Service", "Notes": "Basic library completed."},
# # # #     {"Sno": 8, "Feature": "2D Objects", "Description": "Canvas, Line Renderer, Button", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "UI components completed."},
# # # #     {"Sno": 9, "Feature": "Fire Safety Library", "Description": "Default Fire Safety 3D Models", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Library", "Notes": "Specialized assets completed."},
# # # #     {"Sno": 10, "Feature": "Default Library", "Description": "Common 3D Models (Cube/Capsule)", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Library", "Notes": "Basic primitives completed."},
# # # #     {"Sno": 11, "Feature": "Height Safety Library", "Description": "Default Height Safety 3D Models", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Library", "Notes": "Specialized assets completed."},
# # # #     {"Sno": 12, "Feature": "Hyundai Library", "Description": "Default Hyundai Environment 3D Models", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Library", "Notes": "Specialized assets completed."},
# # # #     {"Sno": 13, "Feature": "Step Manager", "Description": "Order logics with drag and change", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic logic ordering completed."},
# # # #     {"Sno": 14, "Feature": "Add Component", "Description": "Add components if needed", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Core feature completed."},
# # # #     {"Sno": 15, "Feature": "Physics Components", "Description": "Control physics (Rigidbody, Gravity)", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic physics completed."},
# # # #     {"Sno": 16, "Feature": "Colliders", "Description": "Mesh, Box, Sphere, Capsule colliders", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Physics Components", "Notes": "Consolidated colliders."},
# # # #     {"Sno": 17, "Feature": "Render Components", "Description": "Render GameObjects and Meshes", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Consolidated rendering."},
# # # #     {"Sno": 18, "Feature": "Interactables", "Description": "User interactions with GameObjects", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic interactions completed."},
# # # #     {"Sno": 19, "Feature": "Audio", "Description": "Custom audio importation", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Asset Service", "Notes": "Basic audio completed."},
# # # #     {"Sno": 20, "Feature": "HotSpot", "Description": "Show text when hitting hotspot", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Interactables", "Notes": "Basic UI trigger completed."},
# # # #     {"Sno": 21, "Feature": "Attach to Player", "Description": "GameObject follows player", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic tracking completed."},
# # # #     {"Sno": 22, "Feature": "Climb", "Description": "Climbing feature", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Interactables", "Notes": "Basic mechanics completed."},
# # # #     {"Sno": 23, "Feature": "Line Renderer", "Description": "Render lines in scene", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic rendering completed."},
# # # #     {"Sno": 24, "Feature": "Lever", "Description": "Lever interaction feature", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Interactables", "Notes": "Basic mechanics completed."},
# # # #     {"Sno": 25, "Feature": "Knob", "Description": "Knob rotation feature", "Team": "XR", "Status": "Completed", "Priority": "Low", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Interactables", "Notes": "Basic mechanics completed."},
# # # #     {"Sno": 26, "Feature": "Grid", "Description": "Enable/disable grid with snapping", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic snapping completed."},
# # # #     {"Sno": 27, "Feature": "Perspective View", "Description": "Perspective view of scene", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic camera completed."},
# # # #     {"Sno": 28, "Feature": "Orthographic View", "Description": "Orthographic view of scene", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic camera completed."},
# # # #     {"Sno": 29, "Feature": "Child GameObject", "Description": "Custom child objects", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Hierarchy", "Notes": "Core hierarchy completed."},
# # # #     {"Sno": 30, "Feature": "Offline Load/Save", "Description": "Load/save scenes offline", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene Service", "Notes": "Basic offline support completed."},
# # # #     {"Sno": 31, "Feature": "Disable/Enable GO", "Description": "Hide/disable hierarchy objects", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Hierarchy", "Notes": "Basic UI completed."},
# # # #     {"Sno": 32, "Feature": "Drag Drop", "Description": "Drag assets into scene/hierarchy", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Hierarchy", "Notes": "Basic UI completed."},
# # # #     {"Sno": 33, "Feature": "Search", "Description": "Library and project search", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Library", "Notes": "Basic search completed."},
# # # #     {"Sno": 34, "Feature": "Thumbnails", "Description": "Project/library thumbnails", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Library", "Notes": "Basic UI completed."},
# # # #     {"Sno": 35, "Feature": "Copy/Paste Components", "Description": "Copy position/rotation/scale", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic UI completed."},
# # # #     {"Sno": 36, "Feature": "Camera Rotation", "Description": "Rotate camera with mouse", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic camera completed."},
# # # #     {"Sno": 37, "Feature": "Performance Tracking Bar", "Description": "Display mesh usage for JSON", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic UI completed."},
# # # #     {"Sno": 38, "Feature": "Loading Screen", "Description": "Loading UI for login/asset tasks", "Team": "CFT", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Asset Service", "Notes": "Basic UI completed."},
# # # #     {"Sno": 39, "Feature": "Hide/Lock Options", "Description": "Hide Inspector/Hierarchy panels", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic UI completed."},
# # # #     {"Sno": 40, "Feature": "Prevent Multiple Players", "Description": "Limit one player per scene", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic constraint completed."},
# # # #     {"Sno": 41, "Feature": "Sensitivity", "Description": "Mouse control sensitivity", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic input completed."},
# # # #     {"Sno": 42, "Feature": "VR Hand Animation", "Description": "Grip/pinch animations for VR", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Interactables", "Notes": "Basic VR input completed."},
# # # #     {"Sno": 43, "Feature": "Save", "Description": "Explicit Save button in UI", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene Service", "Notes": "Basic UI completed."},
# # # #     {"Sno": 44, "Feature": "Reset", "Description": "Reset transform to zero", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Basic UI completed."}
# # # # ]

# # # # # Create sheets and populate data
# # # # sheets_data = {
# # # #     "Unity Features": unity_features,
# # # #     "Backend Architecture": backend_architecture,
# # # #     "Cross-Platform Requirements": cross_platform,
# # # #     "Product Launch Plan": launch_plan,
# # # #     "Completed Features": completed_features
# # # # }

# # # # # Process each sheet
# # # # for sheet_name, data in sheets_data.items():
# # # #     ws = workbook.create_sheet(title=sheet_name)
# # # #     headers = list(data[0].keys())
    
# # # #     # Write headers
# # # #     for col_idx, header in enumerate(headers, start=1):
# # # #         cell = ws.cell(row=1, column=col_idx)
# # # #         cell.value = header
# # # #         cell.font = Font(bold=True)
# # # #         cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
# # # #     # Sort data by Priority
# # # #     sorted_data = sorted(data, key=lambda x: priority_order.get(x["Priority"], 3))
    
# # # #     # Write data
# # # #     for row_idx, row_data in enumerate(sorted_data, start=2):
# # # #         for col_idx, key in enumerate(headers, start=1):
# # # #             cell = ws.cell(row=row_idx, column=col_idx)
# # # #             cell.value = row_data[key]
# # # #             cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    
# # # #     # Apply filters
# # # #     ws.auto_filter.ref = ws.dimensions
    
# # # #     # Auto-adjust column widths
# # # #     for col_idx in range(1, len(headers) + 1):
# # # #         max_length = 0
# # # #         column_letter = get_column_letter(col_idx)
# # # #         for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
# # # #             for cell in row:
# # # #                 try:
# # # #                     if len(str(cell.value)) > max_length:
# # # #                         max_length = len(str(cell.value))
# # # #                 except:
# # # #                     pass
# # # #         adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap width for readability
# # # #         ws.column_dimensions[column_letter].width = adjusted_width

# # # # # Save the workbook
# # # # workbook.save("VR_Authoring_Tool_Roadmap_Enhanced.xlsx")
# # # # print("Excel file 'VR_Authoring_Tool_Roadmap_Enhanced.xlsx' has been created successfully.")
# # # import openpyxl
# # # from openpyxl.styles import Font, PatternFill
# # # from openpyxl.worksheet.filters import AutoFilter
# # # from openpyxl.utils import get_column_letter

# # # # Create a new workbook
# # # workbook = openpyxl.Workbook()
# # # workbook.remove(workbook.active)

# # # # Define priority order for sorting
# # # priority_order = {"High": 1, "Medium": 2, "Low": 3}

# # # # Data for Frontend Unity Features (client-side VR authoring in Unity)
# # # frontend_unity = [
# # #     {"Feature": "Immersive Scene Editor", "Description": "Real-time VR editing mode for creating/editing scenes in VR.", "Unity Implementation": "XR Interaction Toolkit for hand tracking; VR camera rigs.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Core for intuitive creator experience in Unity editor."},
# # #     {"Feature": "Advanced Asset Import", "Description": "Support GLTF/GLB/FBX/OBJ with auto-optimization.", "Unity Implementation": "AssetPostprocessor; Unity Asset Store API.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Unity-specific import pipeline."},
# # #     {"Feature": "Interactable System", "Description": "Pre-built interactions: grab, throw, UI pointers, haptics.", "Unity Implementation": "XRGrabInteractable, XRSimpleInteractable.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Scene Service", "Notes": "Unity XR toolkit for VR interactions."},
# # #     {"Feature": "Animation & Keyframing", "Description": "Timeline-based animation for objects/characters.", "Unity Implementation": "Unity Timeline & Animator; Mixamo integration.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Asset Service", "Notes": "Unity-native animation tools."},
# # #     {"Feature": "Physics & Simulation", "Description": "Rigidbody, soft-body, fluid sims; raycasting.", "Unity Implementation": "Unity Physics (DOTS); NVIDIA PhysX.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Unity physics engine integration."},
# # #     {"Feature": "Visual Scripting", "Description": "Node-based scripting for no-code behaviors.", "Unity Implementation": "Unity Visual Scripting; custom VR nodes.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Unity's built-in visual scripting."},
# # #     {"Feature": "Real-Time Lighting", "Description": "Global illumination, HDRP/URP, light baking.", "Unity Implementation": "Universal Render Pipeline; adaptive quality.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Unity rendering pipelines."},
# # #     {"Feature": "Hand & Gesture Tracking", "Description": "Hand poses, pinch/grab, finger tracking.", "Unity Implementation": "Oculus Integration; OpenXR.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Interactable System", "Notes": "Unity XR plugins."},
# # #     {"Feature": "Multi-User Collaboration", "Description": "Real-time co-editing with avatars, voice chat.", "Unity Implementation": "Photon Networking; Unity Netcode.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Collaboration Service", "Notes": "Unity networking for VR sessions."},
# # #     {"Feature": "AI-Assisted Creation", "Description": "Auto-generate assets/layouts; smart suggestions.", "Unity Implementation": "ML-Agents; external AI API plugins.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Unity ML integrations."},
# # #     {"Feature": "Preview & Testing Suite", "Description": "In-editor device previews; performance profiling.", "Unity Implementation": "Unity Device Simulator; build pipelines.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Unity editor tools."},
# # #     {"Feature": "Accessibility Features", "Description": "Color-blind modes, subtitles, scalable UI.", "Unity Implementation": "Unity UI Toolkit; accessibility plugins.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "None", "Notes": "Unity UI for VR."},
# # #     {"Feature": "Custom Scripting (C#)", "Description": "IDE integration for advanced scripting.", "Unity Implementation": "Visual Studio; VR script templates.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Visual Scripting", "Notes": "Unity scripting environment."},
# # #     {"Feature": "Audio Spatialization", "Description": "3D audio with occlusion/reverb; MIDI support.", "Unity Implementation": "Unity Audio Spatializer; FMOD.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Asset Service", "Notes": "Unity audio engines."},
# # #     {"Feature": "Particle & Effects System", "Description": "VFX for fire/smoke; custom shaders.", "Unity Implementation": "VFX Graph; Shader Graph.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Real-Time Lighting", "Notes": "Unity graphics tools."}
# # # ]

# # # # Data for Frontend Web Features (web-specific aspects for previews, deployments, WebXR)
# # # frontend_web = [
# # #     {"Feature": "WebXR Support", "Description": "Browser-based VR experiences and previews.", "Implementation": "WebXR API; Unity WebGL exports.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "For web-based viewer app."},
# # #     {"Feature": "Browser Input Handling", "Description": "Mouse/keyboard/touch input for web previews.", "Implementation": "Unity Input System; WebGL input modules.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Fallback for non-VR browsers."},
# # #     {"Feature": "WebGL Rendering Optimization", "Description": "Optimized rendering for browsers (LOD, compression).", "Implementation": "Unity URP for WebGL; dynamic scaling.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Real-Time Lighting", "Notes": "Ensure 60 FPS in Chrome/Firefox."},
# # #     {"Feature": "Asset Compatibility for Web", "Description": "GLTF exports; poly limits for browser performance.", "Implementation": "GLTFast plugin; web preprocessing.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Max 50k polys for web."},
# # #     {"Feature": "Progressive Web App (PWA)", "Description": "Installable web app for offline previews.", "Implementation": "Service workers; Unity WebGL PWA setup.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "For web-based authoring lite."},
# # #     {"Feature": "Web Collaboration", "Description": "Real-time co-editing in browser sessions.", "Implementation": "WebSockets integration with Unity Netcode.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Multi-User Collaboration", "Notes": "Browser avatars and chat."},
# # #     {"Feature": "Web Analytics Integration", "Description": "Track web session usage in browser.", "Implementation": "Google Analytics; Unity web plugins.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Analytics Service", "Notes": "For web viewer feedback."},
# # #     {"Feature": "Web Accessibility", "Description": "WCAG compliance for web UI.", "Implementation": "ARIA labels; keyboard navigation in WebGL.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Accessibility Features", "Notes": "Screen reader support."},
# # #     {"Feature": "Web Notification System", "Description": "Browser push notifications for updates.", "Implementation": "Web Push API; integrated with backend.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Notification Service", "Notes": "For web users."},
# # #     {"Feature": "Web Audio Handling", "Description": "Spatial audio in browsers via WebAudio.", "Implementation": "WebAudio API; fallback for Unity audio.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Audio Spatialization", "Notes": "Cross-browser testing."}
# # # ]

# # # # Data for Backend Features (server-side components)
# # # backend_features = [
# # #     {"Component": "API Gateway", "Description": "Routes requests; handles auth/load balancing.", "Technologies": "AWS API Gateway; OAuth2/JWT.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Low latency for Unity/Web calls."},
# # #     {"Component": "Auth Service", "Description": "User sign-in, RBAC, multi-tenancy.", "Technologies": "Firebase Auth; Keycloak; SSO.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Supports Unity and web logins."},
# # #     {"Component": "Asset Service", "Description": "Upload/process/serve assets; CDN delivery.", "Technologies": "S3/Azure Blob; FFmpeg.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "Optimized for Unity/Web assets."},
# # #     {"Component": "Scene Service", "Description": "CRUD for scenes; version control, autosave.", "Technologies": "MongoDB; Git-like diffs.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "Syncs with Unity/Web editors."},
# # #     {"Component": "Collaboration Service", "Description": "Real-time sync for multi-user editing.", "Technologies": "WebSockets (Socket.io); OT.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Supports Unity/Web sessions."},
# # #     {"Component": "Analytics Service", "Description": "Track usage, errors, engagement; reports.", "Technologies": "Elasticsearch/Kibana.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "API Gateway", "Notes": "Aggregates Unity/Web data."},
# # #     {"Component": "Publishing Service", "Description": "Build/deploy VR apps (APK, WebGL, EXE).", "Technologies": "Jenkins; Unity Cloud Build.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "WebGL builds for web deployment."},
# # #     {"Component": "Database Layer", "Description": "Store users/projects/assets; caching.", "Technologies": "Postgres; MongoDB; Redis.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "Scalable for Unity/Web traffic."},
# # #     {"Component": "AI Integration", "Description": "Backend for AI asset generation.", "Technologies": "Azure OpenAI; Hugging Face.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "API for Unity/Web calls."},
# # #     {"Component": "Security & Compliance", "Description": "Encryption, GDPR, vulnerability scans.", "Technologies": "HTTPS; OWASP practices.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Protects Unity/Web data."},
# # #     {"Component": "Scalability Layer", "Description": "Auto-scaling; load testing.", "Technologies": "Kubernetes; Prometheus.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Handles Unity/Web peaks."},
# # #     {"Component": "Notification Service", "Description": "Push alerts for collaboration/builds.", "Technologies": "Firebase Cloud Messaging.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "API Gateway", "Notes": "For Unity/Web users."}
# # # ]

# # # # Data for Product Launch Plan (high-level, not strictly frontend/backend)
# # # launch_plan = [
# # #     {"Task": "Define Target Audience", "Description": "Identify users (educators, designers, enterprises).", "Details": "Market research; user personas.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Nov 2025", "Dependencies": "None", "Notes": "Covers Unity/Web users."},
# # #     {"Task": "Beta Testing", "Description": "Test with 100+ users; gather feedback.", "Details": "Surveys; beta on Steam/Oculus/web.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Preview & Testing Suite", "Notes": "Test Unity/Web builds."},
# # #     {"Task": "Monetization Strategy", "Description": "Freemium model; paid pro features; subscriptions.", "Details": "Unity Asset Store; tiered pricing.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Auth Service", "Notes": "For Unity/Web access."},
# # #     {"Task": "Compliance & Accessibility", "Description": "GDPR, WCAG, HIPAA; audits.", "Details": "Legal consultation; tools.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Security & Compliance", "Notes": "Applies to Unity/Web."},
# # #     {"Task": "Marketing Campaign", "Description": "Launch on Steam, Oculus, web; social media.", "Details": "X posts; influencer demos.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Promote Unity/Web features."},
# # #     {"Task": "Post-Launch Support", "Description": "Updates, bug fixes, forums.", "Details": "Discord; patch pipeline.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "Support Unity/Web users."},
# # #     {"Task": "Scalability Testing", "Description": "Load test for 1,000 users.", "Details": "AWS testing; monitoring.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scalability Layer", "Notes": "Test Unity/Web traffic."}
# # # ]

# # # # Data for Completed Features (mostly frontend Unity, from original)
# # # completed_features = [
# # #     {"Sno": 1, "Feature": "Hierarchy", "Description": "Displays GameObjects", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Unity frontend completed."},
# # #     {"Sno": 2, "Feature": "Scene", "Description": "Logic Handling Environment", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Unity frontend completed."},
# # #     {"Sno": 3, "Feature": "Move Transform", "Description": "Move a GameObject", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Unity frontend completed."},
# # #     # ... (abbreviated for brevity; include all from previous)
# # #     {"Sno": 44, "Feature": "Reset", "Description": "Reset transform to zero", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Unity frontend completed."}
# # # ]

# # # # Note: For completed_features, I've abbreviated; in full script, copy all 44 from previous.

# # # # Create sheets and populate data
# # # sheets_data = {
# # #     "Frontend Unity": frontend_unity,
# # #     "Frontend Web": frontend_web,
# # #     "Backend": backend_features,
# # #     "Product Launch Plan": launch_plan,
# # #     "Completed Features": completed_features
# # # }

# # # # Process each sheet
# # # for sheet_name, data in sheets_data.items():
# # #     ws = workbook.create_sheet(title=sheet_name)
# # #     headers = list(data[0].keys())
    
# # #     # Write headers
# # #     for col_idx, header in enumerate(headers, start=1):
# # #         cell = ws.cell(row=1, column=col_idx)
# # #         cell.value = header
# # #         cell.font = Font(bold=True)
# # #         cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
# # #     # Sort data by Priority
# # #     sorted_data = sorted(data, key=lambda x: priority_order.get(x["Priority"], 3))
    
# # #     # Write data
# # #     for row_idx, row_data in enumerate(sorted_data, start=2):
# # #         for col_idx, key in enumerate(headers, start=1):
# # #             cell = ws.cell(row=row_idx, column=col_idx)
# # #             cell.value = row_data[key]
# # #             cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    
# # #     # Apply filters
# # #     ws.auto_filter.ref = ws.dimensions
    
# # #     # Auto-adjust column widths
# # #     for col_idx in range(1, len(headers) + 1):
# # #         max_length = 0
# # #         column_letter = get_column_letter(col_idx)
# # #         for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
# # #             for cell in row:
# # #                 try:
# # #                     if len(str(cell.value)) > max_length:
# # #                         max_length = len(str(cell.value))
# # #                 except:
# # #                     pass
# # #         adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap for readability
# # #         ws.column_dimensions[column_letter].width = adjusted_width

# # # # Save the workbook
# # # workbook.save("VR_Authoring_Tool_Bifurcated_Roadmap.xlsx")
# # # print("Excel file 'VR_Authoring_Tool_Bifurcated_Roadmap.xlsx' has been created successfully.")
# # import openpyxl
# # from openpyxl.styles import Font, PatternFill
# # from openpyxl.worksheet.filters import AutoFilter
# # from openpyxl.utils import get_column_letter
# # from datetime import datetime

# # # Create a new workbook
# # workbook = openpyxl.Workbook()
# # workbook.remove(workbook.active)

# # # Define priority order for sorting
# # priority_order = {"High": 1, "Medium": 2, "Low": 3}

# # # Current date for conditional formatting
# # current_date = datetime(2025, 8, 26)

# # # Expanded Data for Frontend Unity Features
# # frontend_unity = [
# #     {"Feature": "Immersive Scene Editor", "Description": "Real-time VR editing environment: Sub-elements include hand gesture controls (pinch, swipe, point), spatial UI panels (resizable, pinnable), multi-select with bounding box, undo/redo with VR gesture triggers, occlusion-aware snapping, and error recovery for large scenes (>10k objects). Workflow: Select object → gesture to move → snap to grid → save state. Metrics: <200ms gesture response, 90 FPS.", "Unity Implementation": "XR Interaction Toolkit v2.5; VR camera rigs with custom input mappings; Unity Undo system with VR events; occlusion culling via URP.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library, Scene Service", "Notes": "Core UX; test for motion sickness (latency <20ms), Quest 3 passthrough support. Risk: Gesture misrecognition."},
# #     {"Feature": "Advanced Asset Import", "Description": "Import GLTF/GLB/FBX/OBJ: Sub-elements include auto-LOD generation (3 levels), texture compression (ASTC/DXT), metadata extraction (tags, materials), batch import with progress UI, error handling (auto-repair corrupt meshes), and validation for VR compatibility. Workflow: Drag file → validate → optimize → add to library. Metrics: <5s for 10MB model.", "Unity Implementation": "AssetPostprocessor with Simplygon integration; Unity Asset Store API; custom importers for VR.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library, Asset Service", "Notes": "Support 100MB+ models; edge cases: malformed FBX, high-poly crashes. Effort: 20 person-days."},
# #     {"Feature": "Interactable System", "Description": "VR interactions: Sub-elements include haptic feedback (variable intensity), multi-object grab (grouping), event triggers for scripts (onClick/onHover), trajectory prediction for throws, voice-activated controls, and collision feedback UI. Workflow: Select object → assign interaction → test in VR. Metrics: <10ms haptic latency.", "Unity Implementation": "XRGrabInteractable with extensions; OpenXR haptics; custom UnityEvent triggers.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Scene Service, Hand & Gesture Tracking", "Notes": "Test on Quest/Vive; edge cases: network lag in multiplayer. Effort: 25 person-days."},
# #     {"Feature": "Animation & Keyframing", "Description": "Timeline animations: Sub-elements include keyframe interpolation (Bezier/slerp), skeletal rigging UI, blend shapes for faces, VR timeline scrubbing (gesture-based), baking for performance, and export to GLTF/USD. Workflow: Select object → add keyframes → preview in VR → export. Metrics: <1s export for 100 keyframes.", "Unity Implementation": "Unity Timeline & Animator; Mixamo auto-rigging; custom VR UI.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Asset Service", "Notes": "Support IK; edge cases: looping jitter. Effort: 30 person-days."},
# #     {"Feature": "Physics & Simulation", "Description": "Advanced physics: Sub-elements include rigidbody joints (hinge/fixed), soft-body (cloth/skin), fluid sims (particle-based), raycasting for queries, collision debug visuals, and performance profiling. Workflow: Add physics component → configure → test collision. Metrics: <5% CPU for 100 rigidbodies.", "Unity Implementation": "Unity Physics (DOTS); NVIDIA PhysX 5.0; Obi Fluid/Cloth.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Optimize for Quest; edge cases: high-velocity crashes. Effort: 35 person-days."},
# #     {"Feature": "Visual Scripting", "Description": "No-code workflows: Sub-elements include VR-optimized node graph (gesture zoom/pan), pre-built VR nodes (onGrab/onTrigger), step-through debugger, node templates (doors, levers), and exportable scripts. Workflow: Drag nodes → connect → test in VR. Metrics: <1s node execution.", "Unity Implementation": "Unity Visual Scripting v1.8; custom VR node UI.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Non-coder focus; edge cases: cyclic graphs. Effort: 20 person-days."},
# #     {"Feature": "Real-Time Lighting", "Description": "Dynamic lighting: Sub-elements include GI probes (real-time/baked), HDRP/URP presets, progressive light baking, shadow mapping (soft/hard), volumetric fog, and adaptive quality sliders. Workflow: Place light → adjust → bake if needed. Metrics: <10s bake for small scene.", "Unity Implementation": "URP v14; Bakery plugin; custom VR light UI.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "90 FPS target; edge cases: dynamic shadows on mobile. Effort: 25 person-days."},
# #     {"Feature": "Hand & Gesture Tracking", "Description": "VR input: Sub-elements include pose recognition (fist/open), pinch precision (mm accuracy), finger joint tracking, gesture combos (swipe/double-tap), controller fallbacks, and calibration UI. Workflow: Calibrate hands → test gestures → assign actions. Metrics: >95% accuracy.", "Unity Implementation": "Oculus Integration v50; OpenXR 1.9; custom recognizers.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Interactable System", "Notes": "Test on Quest 3; edge cases: low-light tracking. Effort: 20 person-days."},
# #     {"Feature": "Multi-User Collaboration", "Description": "Co-editing: Sub-elements include avatar customization (models/colors), spatial voice chat (3D audio), real-time transform sync (<50ms latency), conflict resolution UI, session recordings, and spectator mode. Workflow: Invite user → edit → resolve conflicts. Metrics: 10+ users.", "Unity Implementation": "Photon Fusion; Unity Netcode; Agora for voice.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Collaboration Service", "Notes": "Edge cases: disconnects; effort: 40 person-days."},
# #     {"Feature": "AI-Assisted Creation", "Description": "Smart tools: Sub-elements include procedural terrain/buildings, asset recommendation (context-based), auto-scripting (e.g., door logic), NLP prompts ('add a forest'), ethical filters, and offline mode. Workflow: Input prompt → generate → refine. Metrics: <3s generation.", "Unity Implementation": "ML-Agents v2; Azure OpenAI via backend.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: biased outputs; effort: 50 person-days."},
# #     {"Feature": "Preview & Testing Suite", "Description": "Device testing: Sub-elements include multi-device emulators (Quest/PC/Web), FPS/memory dashboards, automated interaction tests, bug reporting UI, A/B feature testing, and crash analytics. Workflow: Select device → run test → export report. Metrics: <1min test run.", "Unity Implementation": "Unity Device Simulator v3; custom profiling tools.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: device-specific bugs; effort: 30 person-days."},
# #     {"Feature": "Accessibility Features", "Description": "Inclusive design: Sub-elements include color-blind modes, subtitle timing UI, scalable text, voice command parsing (NLP), haptic alternatives, one-handed controls, and audio descriptions. Workflow: Enable setting → test in VR. Metrics: WCAG 2.2 compliance.", "Unity Implementation": "Unity UI Toolkit; Microsoft Seeing AI plugin.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "None", "Notes": "Test with diverse users; effort: 25 person-days."},
# #     {"Feature": "Custom Scripting (C#)", "Description": "Advanced coding: Sub-elements include VR code editor (syntax highlighting), script templates (VR events), breakpoints in VR, Git integration, sandboxed execution, and error consoles. Workflow: Write code → debug → deploy. Metrics: <1s compile.", "Unity Implementation": "Visual Studio Code integration; VR UI overlays.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Visual Scripting", "Notes": "Secure sandbox; edge cases: runtime errors; effort: 30 person-days."},
# #     {"Feature": "Audio Spatialization", "Description": "3D audio: Sub-elements include occlusion filters (material-based), reverb zones, MIDI sequencing UI, volume curves, binaural support, and audio import validation. Workflow: Place audio → configure → test in VR. Metrics: <10ms latency.", "Unity Implementation": "Unity Audio Spatializer; FMOD Studio v2.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Asset Service", "Notes": "Test on Atmos headsets; edge cases: echo; effort: 20 person-days."},
# #     {"Feature": "Particle & Effects System", "Description": "VFX tools: Sub-elements include particle emitters (curves), shader effects (glow/refraction), wind/gravity influencers, collision reactions, preset library (fire/smoke), and VR preview. Workflow: Create effect → adjust → save. Metrics: <5% GPU.", "Unity Implementation": "VFX Graph v14; Shader Graph for URP.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Real-Time Lighting", "Notes": "Optimize mobile; edge cases: overdraw; effort: 25 person-days."},
# #     {"Feature": "Custom Asset Library", "Description": "User-managed assets: Sub-elements include upload queues (async, progress UI), tagging (AI-suggested), thumbnail generation, sharing permissions (ACLs), versioning with diffs, moderation tools, and drag-drop to scenes. Workflow: Upload → tag → share → use in scene. Metrics: <2s thumbnail render.", "Unity Implementation": "Custom UI panels; AssetDatabase for caching; drag-drop scripts.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service, Auth Service", "Notes": "New core feature; edge cases: duplicate tags; effort: 40 person-days."},
# #     {"Feature": "Localization Support", "Description": "Multi-language: Sub-elements include metadata translation, UI string keys, locale detection, RTL support, export to .po files, and translation memory. Workflow: Set locale → update UI → export. Metrics: <1s UI switch.", "Unity Implementation": "Unity Localization v1.4; custom VR UI adapters.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "Support 20 languages; edge cases: non-Latin; effort: 20 person-days."},
# #     {"Feature": "Performance Monitoring", "Description": "In-editor tools: Sub-elements include FPS graphs (real-time), memory trackers, draw call analyzers, VR-specific metrics (ASW usage), auto-optimization suggestions, and exportable reports. Workflow: Open profiler → analyze → optimize. Metrics: <1% overhead.", "Unity Implementation": "Unity Profiler extensions; custom VR dashboards.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Preview & Testing Suite", "Notes": "Edge cases: VR stutter; effort: 25 person-days."},
# #     {"Feature": "User Feedback Loops", "Description": "In-app feedback: Sub-elements include survey popups, error auto-reporting, feature voting UI, session replay for bugs, and integration with analytics. Workflow: Encounter issue → report → track resolution. Metrics: <10s report submission.", "Unity Implementation": "Unity Analytics; custom UI modals.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "GDPR-compliant; effort: 15 person-days."},
# #     {"Feature": "VR-Specific Debugging", "Description": "In-VR tools: Sub-elements include heatmap overlays for interactions, breakpoint triggers in VR, console logs in spatial UI, performance alerts, and remote debugging endpoints. Workflow: Enter debug mode → set breakpoint → inspect. Metrics: <2% FPS drop.", "Unity Implementation": "Custom VR debug UI; Unity Debug.Log extensions.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Custom Scripting", "Notes": "New; edge cases: debug UI occlusion; effort: 25 person-days."},
# #     {"Feature": "AR Support", "Description": "Mobile AR authoring: Sub-elements include ARKit/ARCore integration, plane detection, anchor placement, AR preview in VR, and export to AR formats. Workflow: Scan environment → place asset → test. Metrics: <1s plane detection.", "Unity Implementation": "AR Foundation v5; XR Management for multi-platform.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "New; test on iOS/Android; effort: 30 person-days."},
# #     {"Feature": "Asset Marketplace", "Description": "User-contributed assets: Sub-elements include upload moderation, ratings/reviews UI, search by popularity, payment integration (e.g., Stripe), and analytics on sales. Workflow: Submit asset → review → publish → purchase. Metrics: <5s search.", "Unity Implementation": "Custom UI; Unity Asset Store API hooks.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "New; edge cases: fraud detection; effort: 40 person-days."},
# #     {"Feature": "Dynamic Tutorials", "Description": "In-app XR guides: Sub-elements include voiceover guides, interactive steps, localized scripts, progress tracking, and branching tutorials based on skill. Workflow: Start tutorial → follow VR prompts → complete. Metrics: <1min load.", "Unity Implementation": "Custom UI overlays; Unity Addressables for streaming.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Localization Support", "Notes": "New; edge cases: user interruptions; effort: 20 person-days."},
# #     {"Feature": "Session Recovery", "Description": "Crash recovery: Sub-elements include auto-save every 15s, restore UI with diff highlights, cloud sync for cross-device, error logs, and user prompts for recovery. Workflow: Crash → restart → recover. Metrics: <5s restore.", "Unity Implementation": "Custom autosave scripts; Unity Cloud sync.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scene Service", "Notes": "New; edge cases: partial saves; effort: 15 person-days."}
# # ]

# # # Expanded Data for Frontend Web Features
# # frontend_web = [
# #     {"Feature": "WebXR Support", "Description": "Browser VR: Sub-elements include VR session initiation (button UX), immersive mode toggles (VR/2D), device permission dialogs, fallback to 2D with orbit controls, progressive enhancement for mobile, and WebXR hit-test for AR. Workflow: Click enter VR → detect headset → render scene. Metrics: <2s session start.", "Implementation": "WebXR API 1.1; Unity WebGL v2023; A-Frame fallbacks.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Test Chrome/Edge/Safari; edge cases: permission denials; effort: 25 person-days."},
# #     {"Feature": "Browser Input Handling", "Description": "Web inputs: Sub-elements include touch gestures (pinch-zoom), keyboard shortcuts (WASD), mouse orbit with inertia, gamepad support (Xbox/PS), sensitivity sliders, and multi-touch detection. Workflow: Select input → interact → save settings. Metrics: <10ms input latency.", "Implementation": "Unity Input System; WebGL Pointer Lock; Hammer.js for touch.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "iOS Safari compatibility; edge cases: touch conflicts; effort: 20 person-days."},
# #     {"Feature": "WebGL Rendering Optimization", "Description": "Browser rendering: Sub-elements include dynamic resolution scaling, texture mipmapping (4 levels), shader fallbacks (GLES2/3), progressive loading bars, context loss recovery, and GPU memory alerts. Workflow: Load scene → adjust quality → monitor FPS. Metrics: 60 FPS on mid-tier PCs.", "Implementation": "Unity URP v14; custom WebGL scripts.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Real-Time Lighting", "Notes": "Test on Intel GPUs; edge cases: context loss; effort: 25 person-days."},
# #     {"Feature": "Asset Compatibility for Web", "Description": "Web assets: Sub-elements include GLTF validation (schema checks), poly reduction (target 50k), streaming asset loads, format converters (OBJ to GLTF), cache eviction policies, and CORS handlers. Workflow: Upload asset → validate → stream. Metrics: <3s for 10MB GLTF.", "Implementation": "GLTFast v5; Web Workers for preprocessing.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Edge cases: CORS errors; effort: 20 person-days."},
# #     {"Feature": "Progressive Web App (PWA)", "Description": "Installable app: Sub-elements include manifest configs (icons, themes), offline asset caching, service worker updates, push notification hooks, homescreen shortcuts, and install prompts. Workflow: Visit site → install → use offline. Metrics: <5MB offline cache.", "Implementation": "Workbox for service workers; Unity WebGL PWA.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: storage limits; effort: 15 person-days."},
# #     {"Feature": "Web Collaboration", "Description": "Browser co-editing: Sub-elements include cursor sharing (colored pointers), chat overlays (text/emoji), version locking, invite links with expiry, tab sync, and conflict merge UI. Workflow: Invite → edit → resolve conflicts. Metrics: <50ms sync latency.", "Implementation": "WebSockets; Firebase Realtime DB; Yjs for OT.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Multi-User Collaboration", "Notes": "Support 5 users; edge cases: tab crashes; effort: 30 person-days."},
# #     {"Feature": "Web Analytics Integration", "Description": "Session tracking: Sub-elements include event funnels (click-to-publish), heatmaps for UI, A/B test APIs, privacy opt-outs, CSV exports, and real-time dashboards. Workflow: Track session → generate report → optimize. Metrics: <1s event log.", "Implementation": "Google Analytics 4; custom WebGL events.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Analytics Service", "Notes": "GDPR-compliant; edge cases: ad blockers; effort: 15 person-days."},
# #     {"Feature": "Web Accessibility", "Description": "Inclusive UI: Sub-elements include ARIA labels for 3D objects, keyboard navigation (tab-index), screen reader support, contrast checkers, reduced motion modes, and font scaling. Workflow: Enable accessibility → navigate → test. Metrics: WCAG 2.2 AA.", "Implementation": "ARIA standards; WebGL accessibility polyfills.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Accessibility Features", "Notes": "Test with NVDA; effort: 20 person-days."},
# #     {"Feature": "Web Notification System", "Description": "Browser alerts: Sub-elements include permission dialogs, customizable channels, vibration for mobile, badge updates, offline queuing, and delivery analytics. Workflow: Subscribe → receive alert → act. Metrics: <2s delivery.", "Implementation": "Web Push API; Service Workers.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Notification Service", "Notes": "Edge cases: denied permissions; effort: 10 person-days."},
# #     {"Feature": "Web Audio Handling", "Description": "Browser sound: Sub-elements include WebAudio spatial nodes, context resumption, format fallbacks (MP3/Ogg), volume normalization, background playback, and audio debug tools. Workflow: Load audio → configure → test. Metrics: <10ms latency.", "Implementation": "WebAudio API; Unity audio fallbacks.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Audio Spatialization", "Notes": "Safari fixes; effort: 15 person-days."},
# #     {"Feature": "Custom Asset Library (Web)", "Description": "Web dashboard: Sub-elements include drag-drop uploads, WebGL previews (360° view), metadata editor (tags/categories), sharing via secure links, collaborative editing, and export to Unity. Workflow: Upload → tag → share → import to Unity. Metrics: <3s preview render.", "Implementation": "React 18 with Tailwind; WebGL preview shaders.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library (Backend)", "Notes": "Mobile-responsive; edge cases: large uploads; effort: 30 person-days."},
# #     {"Feature": "Web Localization", "Description": "Browser multi-language: Sub-elements include browser locale detection, dynamic string loading, RTL layouts, currency/date formatting, translation API hooks (e.g., DeepL), and caching. Workflow: Detect locale → load UI → export. Metrics: <1s switch.", "Implementation": "react-i18next; WebGL text rendering.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Localization Support", "Notes": "Support 15 languages; effort: 15 person-days."},
# #     {"Feature": "Web Performance Tools", "Description": "Browser monitoring: Sub-elements include Lighthouse audits, network throttle simulators, asset load timelines, WebGL error consoles, optimization suggestions, and exportable reports. Workflow: Run audit → analyze → optimize. Metrics: <1% overhead.", "Implementation": "Chrome DevTools APIs; custom WebGL profilers.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Performance Monitoring", "Notes": "Edge cases: slow 3G; effort: 20 person-days."},
# #     {"Feature": "Web AR Support", "Description": "Browser AR: Sub-elements include WebXR AR mode, plane detection via hit-test, anchor placement UI, mobile camera passthrough, and export to AR formats. Workflow: Scan room → place asset → test. Metrics: <1s plane detection.", "Implementation": "WebXR AR API; 8th Wall fallback.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "AR Support", "Notes": "New; test on iOS/Android; effort: 25 person-days."},
# #     {"Feature": "Web Asset Marketplace", "Description": "Browser asset store: Sub-elements include moderated uploads, ratings/reviews UI, search by tags/ratings, Stripe payments, and analytics dashboards. Workflow: Browse → purchase → import. Metrics: <5s search.", "Implementation": "React SPA; Stripe SDK.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Asset Marketplace", "Notes": "New; edge cases: payment failures; effort: 30 person-days."}
# # ]

# # # Expanded Data for Backend Features
# # backend_features = [
# #     {"Component": "API Gateway", "Description": "Routing hub: Sub-elements include rate limiting (100 req/s/user), API versioning (v1/v2), audit logging (request/response), CORS for Unity/Web, health check endpoints, and throttling for DDoS protection. Endpoints: GET /health, POST /auth/token. Metrics: <100ms latency.", "Technologies": "AWS API Gateway; Kong; OAuth2/JWT with Redis cache.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "99.99% uptime; edge cases: traffic spikes; effort: 30 person-days."},
# #     {"Component": "Auth Service", "Description": "User management: Sub-elements include passwordless login (email links), 2FA (TOTP/SMS), RBAC (admin/creator/viewer), session revocation API, audit trails, and SSO (Google/Microsoft). Endpoints: POST /login, GET /users/me. Metrics: <500ms auth.", "Technologies": "Firebase Auth; Keycloak; PostgreSQL for users.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "GDPR data export; edge cases: token theft; effort: 25 person-days."},
# #     {"Component": "Asset Service", "Description": "Asset handling: Sub-elements include virus scanning (ClamAV), metadata indexing (tags/size), thumbnail generation (ImageMagick), deduplication by hash, purge APIs, and CDN caching. Endpoints: POST /assets/upload, GET /assets/{id}. Metrics: <5s for 100MB.", "Technologies": "S3; DynamoDB; FFmpeg; Cloudflare CDN.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "1GB+ files; edge cases: corrupt uploads; effort: 35 person-days."},
# #     {"Component": "Scene Service", "Description": "Scene management: Sub-elements include JSON schema validation, version diffs (delta encoding), autosave (every 15s), rollback endpoints, USD export, and conflict detection. Endpoints: POST /scenes, GET /scenes/{id}/versions. Metrics: <1s save.", "Technologies": "MongoDB 6; libgit2 for diffs; Redis for locks.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "10MB+ scenes; edge cases: merge conflicts; effort: 30 person-days."},
# #     {"Component": "Collaboration Service", "Description": "Real-time sync: Sub-elements include OT for transforms, presence indicators (user cursors), chat persistence, bandwidth throttling (500kbps/user), offline queues, and session analytics. Endpoints: WS /collab/join. Metrics: <50ms sync.", "Technologies": "Socket.io; Yjs; Redis Pub/Sub.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "50 users/session; edge cases: disconnects; effort: 40 person-days."},
# #     {"Component": "Analytics Service", "Description": "Metrics tracking: Sub-elements include funnels (import-to-publish), cohort analysis, error aggregation (stack traces), real-time Kibana dashboards, CSV/JSON exports, and GDPR opt-outs. Endpoints: POST /events, GET /reports. Metrics: <1s log.", "Technologies": "Elasticsearch 8; Kafka; Kibana.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "API Gateway", "Notes": "Anonymized data; edge cases: large datasets; effort: 25 person-days."},
# #     {"Component": "Publishing Service", "Description": "Build/deployment: Sub-elements include CI/CD pipelines (APK/WebGL/EXE), build queues, version tagging, rollback APIs, build analytics, and platform-specific configs. Endpoints: POST /builds, GET /builds/{id}. Metrics: <5min build.", "Technologies": "Jenkins; Unity Cloud Build; Docker.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: build failures; effort: 30 person-days."},
# #     {"Component": "Database Layer", "Description": "Data storage: Sub-elements include sharding (user/scenes), backup scripts (daily), index optimization, Redis caching (1hr TTL), schema migration tools, and high-availability configs. Schema: users, scenes, assets. Metrics: <10ms query.", "Technologies": "Postgres 15; MongoDB 6; Redis 7.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "99.9% uptime; edge cases: corruption; effort: 35 person-days."},
# #     {"Component": "AI Integration", "Description": "AI backend: Sub-elements include prompt templating, rate limiting (10 req/min), caching generated assets, model selection (GPT-4o/Stable Diffusion), billing trackers, and error handling. Endpoints: POST /ai/generate. Metrics: <3s response.", "Technologies": "Azure OpenAI; Hugging Face; SQS queues.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Cost control; edge cases: API downtime; effort: 40 person-days."},
# #     {"Component": "Security & Compliance", "Description": "Protection: Sub-elements include AES-256 encryption, OWASP ZAP scans, GDPR/CCPA/HIPAA reports, intrusion detection (WAF), audit logs, and data anonymization. Endpoints: GET /compliance/report. Metrics: 0 vulnerabilities.", "Technologies": "HTTPS/TLS 1.3; OWASP tools; CloudTrail.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Quarterly pentests; edge cases: DDoS; effort: 30 person-days."},
# #     {"Component": "Scalability Layer", "Description": "Auto-scaling: Sub-elements include horizontal pod scaling, load balancer configs, stress test scripts (Locust), metrics exporters (CPU/RAM), cost dashboards, and failover scripts. Metrics: 10k concurrent users.", "Technologies": "Kubernetes 1.27; Prometheus; Grafana.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: spike traffic; effort: 35 person-days."},
# #     {"Component": "Notification Service", "Description": "Alerts: Sub-elements include template rendering (email/push), subscription UI, retry logic (3 attempts), analytics (open rates), SMS integration (Twilio), and spam filters. Endpoints: POST /notifications. Metrics: <2s delivery.", "Technologies": "Firebase Cloud Messaging; AWS SES; Twilio.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: undeliverable; effort: 15 person-days."},
# #     {"Component": "Custom Asset Library (Backend)", "Description": "Asset storage: Sub-elements include ACLs for permissions, versioning (delta diffs), Elasticsearch indexing, auto-cleanup of orphans, bulk operation APIs, moderation workflows, and audit logs. Endpoints: POST /library/upload, GET /library/search. Metrics: <500ms search.", "Technologies": "S3; DynamoDB; Elasticsearch 8.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service, Auth Service", "Notes": "Scale to 100TB; edge cases: access conflicts; effort: 40 person-days."},
# #     {"Component": "Session Recovery Service", "Description": "Crash recovery: Sub-elements include autosave triggers (every 10s), diff-based restore, cross-device sync, error logging, and recovery UI endpoints. Endpoints: POST /sessions/restore. Metrics: <5s restore.", "Technologies": "MongoDB; Redis; S3 for logs.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scene Service", "Notes": "New; edge cases: partial saves; effort: 20 person-days."},
# #     {"Component": "Asset Marketplace Service", "Description": "Market backend: Sub-elements include moderation queues, payment processing (Stripe), royalty tracking, search indexing, analytics dashboards, and fraud detection. Endpoints: POST /market/submit, GET /market/assets. Metrics: <5s search.", "Technologies": "Stripe; Elasticsearch; DynamoDB.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "New; edge cases: payment disputes; effort: 35 person-days."}
# # ]

# # # Expanded Data for Product Launch Plan
# # launch_plan = [
# #     {"Task": "Define Target Audience", "Description": "Market research: Sub-elements include user surveys (needs analysis), persona creation (indie dev/enterprise), competitor benchmarking (Unity Editor, Adobe Aero), platform segmentation (Unity/Web), and use case validation (training/gaming). Workflow: Survey → analyze → define segments. Metrics: 500+ responses.", "Details": "SurveyMonkey; X polls; industry reports.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Nov 2025", "Dependencies": "None", "Notes": "Focus on XR training; edge cases: niche markets; effort: 20 person-days."},
# #     {"Task": "Beta Testing", "Description": "User trials: Sub-elements include alpha (internal, 20 users), public beta (200 users), feedback forms (Likert scales), bug triage (Jira), A/B feature tests, and diversity testing (devices/regions). Workflow: Deploy beta → collect feedback → iterate. Metrics: <5% crash rate.", "Details": "TestFlight; Steam Beta; Oculus App Lab.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Preview & Testing Suite", "Notes": "Weekly iterations; edge cases: low-end devices; effort: 40 person-days."},
# #     {"Task": "Monetization Strategy", "Description": "Revenue models: Sub-elements include freemium (core free, pro $20/month), in-app asset purchases ($1-10), enterprise licenses ($5k/year), affiliate links (Unity Store), A/B pricing tests, and churn analytics. Workflow: Set pricing → test → launch. Metrics: 10% conversion.", "Details": "Stripe; PayPal; analytics for conversions.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Auth Service, Asset Marketplace", "Notes": "Edge cases: refund disputes; effort: 30 person-days."},
# #     {"Task": "Compliance & Accessibility", "Description": "Legal standards: Sub-elements include GDPR data mapping, WCAG 2.2 audits (WAVE), HIPAA for health sims, CCPA compliance, privacy policy generator, and third-party certifications. Workflow: Audit → remediate → certify. Metrics: 0 non-compliance issues.", "Details": "Legal reviews; automated tools like Axe.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Security & Compliance", "Notes": "Annual audits; edge cases: new regulations; effort: 25 person-days."},
# #     {"Task": "Marketing Campaign", "Description": "Promotion: Sub-elements include X/Reddit teasers, YouTube demo videos, influencer partnerships (VR streamers), email campaigns (Mailchimp), virtual launch events, and ad budgets ($10k). Workflow: Plan → execute → track KPIs. Metrics: 1M impressions.", "Details": "Hootsuite; Google Ads; VR conventions.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Track ROI; edge cases: negative feedback; effort: 35 person-days."},
# #     {"Task": "Post-Launch Support", "Description": "Maintenance: Sub-elements include monthly patches, Discord forums, user guides (video/text), churn analysis (5% target), feature voting portals, and roadmap updates. Workflow: Monitor issues → patch → communicate. Metrics: <24hr response time.", "Details": "GitHub Issues; Zendesk; patch pipelines.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "Edge cases: breaking updates; effort: 30 person-days."},
# #     {"Task": "Scalability Testing", "Description": "Load simulations: Sub-elements include user spike tests (10k users), DB stress tests, failover drills, cost projections ($5k/month), optimization scripts, and monitoring dashboards. Workflow: Simulate → analyze → optimize. Metrics: 99.99% uptime.", "Details": "Locust; AWS Load Testing; New Relic.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scalability Layer", "Notes": "Edge cases: regional outages; effort: 25 person-days."},
# #     {"Task": "User Onboarding", "Description": "User education: Sub-elements include interactive VR tutorials, onboarding emails, progress tracking, gamified milestones, and analytics on completion rates. Workflow: New user → complete tutorial → start project. Metrics: 80% completion.", "Details": "Custom UI; SendGrid for emails.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Dynamic Tutorials", "Notes": "New; edge cases: drop-offs; effort: 20 person-days."},
# #     {"Task": "Community Building", "Description": "Engage users: Sub-elements include Discord server setup, monthly AMAs, user showcases (featured projects), asset contests, and feedback integration. Workflow: Launch server → host events → iterate. Metrics: 1k active members.", "Details": "Discord; X community posts.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "User Feedback Loops", "Notes": "Edge cases: toxic behavior; effort: 15 person-days."}
# # ]

# # # Expanded Data for Completed Features (original with sub-elements)
# # completed_features = [
# #     {"Sno": 1, "Feature": "Hierarchy", "Description": "GameObject management: Sub-elements include tree view (expand/collapse), search by name/type, drag-reorder, multi-select delete, and parent-child linking UI. Workflow: Open hierarchy → reorder → save. Metrics: <100ms refresh.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Basic Unity UI; integrate with Custom Asset Library; effort: 10 person-days."},
# #     {"Sno": 2, "Feature": "Scene", "Description": "Environment logic: Sub-elements include scene loading (progress bar), multi-scene support, layer management, prefab instantiation, and save/load UI. Workflow: Create scene → add objects → save. Metrics: <2s load.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Add multi-user sync; effort: 15 person-days."},
# #     {"Sno": 3, "Feature": "Move Transform", "Description": "Object movement: Sub-elements include gizmo controls, numeric input, snap-to-grid, multi-object move, and undo support. Workflow: Select → move → confirm. Metrics: <10ms response.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Add VR gesture support; effort: 5 person-days."},
# #     # ... (include all 44; abbreviated here for brevity)
# #     {"Sno": 44, "Feature": "Reset", "Description": "Transform reset: Sub-elements include one-click zeroing (position/rotation/scale), batch reset, undo integration, confirmation dialogs, and gizmo reset visuals. Workflow: Select → reset → confirm. Metrics: <100ms action.", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "Scene", "Notes": "Add VR button; effort: 5 person-days."}
# # ]

# # # Expanded Sub-Features Breakdown (5-10 per major feature)
# # sub_features = [
# #     {"Parent Feature": "Immersive Scene Editor", "Sub-Feature": "Hand Gesture Controls", "Description": "Pinch-to-select, swipe for menus; includes calibration UI, gesture combos (e.g., double-tap), and sensitivity sliders. Metrics: >95% accuracy.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Hand & Gesture Tracking", "Notes": "Edge cases: hand occlusion; effort: 10 person-days."},
# #     {"Parent Feature": "Immersive Scene Editor", "Sub-Feature": "Undo/Redo Stack", "Description": "Spatial UI timeline; supports 100 actions, gesture triggers, and diff previews. Workflow: Gesture undo → preview change → confirm. Metrics: <200ms response.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Scene Service", "Notes": "Edge cases: large undo stacks; effort: 8 person-days."},
# #     {"Parent Feature": "Immersive Scene Editor", "Sub-Feature": "Spatial UI Panels", "Description": "Resizable, pinnable panels for properties; supports drag, transparency, and multi-panel layouts. Metrics: <1s refresh.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: panel overlap; effort: 10 person-days."},
# #     {"Parent Feature": "Immersive Scene Editor", "Sub-Feature": "Multi-Select Tools", "Description": "Bounding box selection; supports batch transforms, grouping, and VR lasso gestures. Metrics: <10ms selection.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Edge cases: large object counts; effort: 8 person-days."},
# #     {"Parent Feature": "Immersive Scene Editor", "Sub-Feature": "Occlusion Snapping", "Description": "Snap objects to surfaces with occlusion awareness; includes grid toggle and precision sliders. Metrics: <5ms snap.", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Physics & Simulation", "Notes": "Edge cases: complex geometries; effort: 7 person-days."},
# #     {"Parent Feature": "Advanced Asset Import", "Sub-Feature": "Auto-LOD Generation", "Description": "Creates low/med/high poly versions; supports poly targets (1k-50k), auto-decimation, and preview UI. Metrics: <10s for 10k polys.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Edge cases: malformed meshes; effort: 10 person-days."},
# #     {"Parent Feature": "Advanced Asset Import", "Sub-Feature": "Texture Compression", "Description": "ASTC/DXT compression; supports 4K textures, mipmap generation, and quality previews. Metrics: <5s for 4K texture.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Edge cases: alpha channel issues; effort: 8 person-days."},
# #     {"Parent Feature": "Advanced Asset Import", "Sub-Feature": "Metadata Extraction", "Description": "Pulls tags, materials, animations; supports FBX/OBJ headers, user-editable fields. Metrics: <1s extraction.", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library", "Notes": "Edge cases: missing metadata; effort: 6 person-days."},
# #     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Asset Tagging System", "Description": "User-defined tags; AI-suggested tags via NLP; supports hierarchical categories, search facets. Metrics: <500ms search.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "AI Integration", "Notes": "Edge cases: duplicate tags; effort: 10 person-days."},
# #     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Version History", "Description": "Tracks changes with diffs; supports rollback, compare UI, and version comments. Metrics: <1s diff load.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Edge cases: merge conflicts; effort: 8 person-days."},
# #     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Sharing Permissions", "Description": "ACLs for public/private/team; includes invite links, role-based access, and audit logs. Metrics: <200ms permission check.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Auth Service", "Notes": "Edge cases: revoked access; effort: 10 person-days."},
# #     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Thumbnail Generation", "Description": "Auto-renders previews for assets; supports 360° views, customizable angles, and batch processing. Metrics: <2s render.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Edge cases: high-poly previews; effort: 8 person-days."},
# #     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Moderation Tools", "Description": "Flag/review inappropriate content; includes reporting UI, admin dashboards, and auto-flagging via AI. Metrics: <1s flag processing.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: false positives; effort: 7 person-days."},
# #     {"Parent Feature": "Multi-User Collaboration", "Sub-Feature": "Avatar Customization", "Description": "User-selectable models, colors, skins; persistent profiles with sync. Metrics: <1s avatar load.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Auth Service", "Notes": "Edge cases: low-poly rendering; effort: 8 person-days."},
# #     {"Parent Feature": "Multi-User Collaboration", "Sub-Feature": "Spatial Voice Chat", "Description": "3D audio with attenuation; includes mute, volume sliders, and echo cancellation. Metrics: <20ms latency.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Audio Spatialization", "Notes": "Edge cases: background noise; effort: 8 person-days."},
# #     {"Parent Feature": "AI-Assisted Creation", "Sub-Feature": "Procedural Generation", "Description": "Auto-builds terrain/buildings from seeds; customizable parameters (density/scale). Metrics: <5s generation.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: unrealistic outputs; effort: 10 person-days."},
# #     {"Parent Feature": "AI-Assisted Creation", "Sub-Feature": "Natural Language Prompts", "Description": "Text-to-asset (e.g., 'create castle'); supports ambiguity resolution, style selection. Metrics: <3s response.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: vague prompts; effort: 10 person-days."},
# #     {"Parent Feature": "WebXR Support", "Sub-Feature": "Session Initiation", "Description": "Enter VR button with device detection; includes permission dialogs, error messages. Metrics: <2s start.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: unsupported browsers; effort: 5 person-days."},
# #     {"Parent Feature": "WebXR Support", "Sub-Feature": "Immersive Mode Toggles", "Description": "Switch VR/2D modes; includes UI animations, fallback controls. Metrics: <1s toggle.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Browser Input Handling", "Notes": "Edge cases: mode conflicts; effort: 5 person-days."}
# #     # Add ~100 more sub-features for all major features (e.g., 5-10 per parent)
# # ]

# # # Create sheets and populate data
# # sheets_data = {
# #     "Frontend Unity": frontend_unity,
# #     "Frontend Web": frontend_web,
# #     "Backend": backend_features,
# #     "Product Launch Plan": launch_plan,
# #     "Completed Features": completed_features,
# #     "Sub-Features Breakdown": sub_features
# # }

# # # Process each sheet
# # for sheet_name, data in sheets_data.items():
# #     ws = workbook.create_sheet(title=sheet_name)
# #     headers = list(data[0].keys())
    
# #     # Write headers
# #     for col_idx, header in enumerate(headers, start=1):
# #         cell = ws.cell(row=1, column=col_idx)
# #         cell.value = header
# #         cell.font = Font(bold=True)
# #         cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
# #     # Sort data by Priority
# #     sorted_data = sorted(data, key=lambda x: priority_order.get(x["Priority"], 3))
    
# #     # Write data
# #     for row_idx, row_data in enumerate(sorted_data, start=2):
# #         for col_idx, key in enumerate(headers, start=1):
# #             cell = ws.cell(row=row_idx, column=col_idx)
# #             cell.value = row_data[key]
# #             cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
# #             # Conditional formatting for overdue tasks
# #             timeline_str = row_data.get("Timeline", "")
# #             try:
# #                 timeline_date = datetime.strptime(timeline_str, "%b %Y")
# #                 if timeline_date < current_date and row_data.get("Status", "") != "Completed":
# #                     cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
# #             except:
# #                 pass
    
# #     # Apply filters
# #     ws.auto_filter.ref = ws.dimensions
    
# #     # Auto-adjust column widths (increased cap for extensive descriptions)
# #     for col_idx in range(1, len(headers) + 1):
# #         max_length = 0
# #         column_letter = get_column_letter(col_idx)
# #         for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
# #             for cell in row:
# #                 try:
# #                     if len(str(cell.value)) > max_length:
# #                         max_length = len(str(cell.value))
# #                 except:
# #                     pass
# #         adjusted_width = min((max_length + 2) * 1.2, 70)  # Cap at 70 for readability
# #         ws.column_dimensions[column_letter].width = adjusted_width

# # # Save the workbook
# # workbook.save("VR_Authoring_Tool_Bifurcated_Roadmap_Extensive.xlsx")
# # print("Excel file 'VR_Authoring_Tool_Bifurcated_Roadmap_Extensive.xlsx' has been created successfully.")
# import openpyxl
# from openpyxl.styles import Font, PatternFill
# from openpyxl.worksheet.filters import AutoFilter
# from openpyxl.utils import get_column_letter
# from datetime import datetime

# # Create a new workbook
# workbook = openpyxl.Workbook()
# workbook.remove(workbook.active)

# # Define priority order for sorting
# priority_order = {"High": 1, "Medium": 2, "Low": 3}

# # Current date for conditional formatting
# current_date = datetime(2025, 8, 26)

# # Frontend Unity Features (focused on VR app development)
# frontend_unity = [
#     {"Feature": "App Logic Editor", "Description": "Visual editor for VR app logic: Sub-elements include state machine UI (drag-drop states), event triggers (onInput/onEvent), cross-platform input mappings (Oculus/SteamVR), scriptable logic templates, and error checking for state conflicts. Workflow: Create state → link events → test on platform. Metrics: <1s state transition.", "Unity Implementation": "XR Interaction Toolkit v2.5; Unity Visual Scripting for states; OpenXR input mappings.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library, Scene Service", "Notes": "Core for app devs; test Quest 3/SteamVR; edge cases: input conflicts; effort: 30 person-days."},
#     {"Feature": "Multi-Platform Build Tools", "Description": "Build VR apps for Quest/PC/WebXR: Sub-elements include build config UI (APK/EXE/WebGL), automated SDK injection (Oculus/SteamVR), platform-specific optimizations (e.g., Quest LODs), build validation, and error logs. Workflow: Select platform → configure → build. Metrics: <5min build.", "Unity Implementation": "Unity Build Pipeline; custom SDK scripts; Unity Cloud Build.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Edge cases: SDK version mismatches; effort: 35 person-days."},
#     {"Feature": "Interactable System", "Description": "App interactions: Sub-elements include haptic feedback (variable intensity), multi-object grab for app controls, platform-specific input bindings, event triggers for app logic, and voice input support. Workflow: Assign interaction → test on platform → save. Metrics: <10ms haptic latency.", "Unity Implementation": "XRGrabInteractable; OpenXR haptics; custom UnityEvent system.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Scene Service", "Notes": "Test on Quest/Vive; edge cases: multi-platform input; effort: 25 person-days."},
#     {"Feature": "Animation for Apps", "Description": "App animations: Sub-elements include keyframe UI for UI transitions, skeletal rigging for characters, blend shapes for app states, VR preview for animations, and export to platform formats (GLTF/USD). Workflow: Create animation → test → export. Metrics: <1s export.", "Unity Implementation": "Unity Timeline v1.8; Mixamo; custom VR UI.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: platform-specific playback; effort: 30 person-days."},
#     {"Feature": "Physics for Apps", "Description": "App physics: Sub-elements include rigidbody for interactive controls, soft-body for UI effects, collision events for app logic, raycasting for selections, and platform-optimized physics. Workflow: Add physics → configure → test. Metrics: <5% CPU.", "Unity Implementation": "Unity Physics (DOTS); NVIDIA PhysX 5.0.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Optimize for Quest; edge cases: high-velocity; effort: 30 person-days."},
#     {"Feature": "Visual Scripting for Apps", "Description": "No-code app logic: Sub-elements include node-based UI for app flows, VR event nodes (onInput/onStateChange), debugging tools, templates for app patterns (menus, inventory), and exportable scripts. Workflow: Create nodes → test → deploy. Metrics: <1s execution.", "Unity Implementation": "Unity Visual Scripting v1.8; custom VR nodes.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: cyclic graphs; effort: 20 person-days."},
#     {"Feature": "App Rendering", "Description": "Cross-platform rendering: Sub-elements include URP/HDRP toggles, platform-specific shaders (Quest/PC), light baking for apps, volumetric effects for UI, and adaptive quality. Workflow: Configure render → test platform → optimize. Metrics: 90 FPS.", "Unity Implementation": "URP v14; Bakery plugin; custom VR shaders.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: mobile performance; effort: 25 person-days."},
#     {"Feature": "Input Configurator", "Description": "Platform input mappings: Sub-elements include gesture recognition (pinch/swipe), controller bindings (Oculus/SteamVR), keyboard/gamepad support, input remapping UI, and calibration tools. Workflow: Map input → test → save. Metrics: >95% accuracy.", "Unity Implementation": "OpenXR 1.9; Oculus Integration v50.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Interactable System", "Notes": "Edge cases: input conflicts; effort: 20 person-days."},
#     {"Feature": "Multi-User App Collaboration", "Description": "Co-develop apps: Sub-elements include avatar systems, spatial voice chat, real-time code sync (<50ms), conflict resolution UI, and session recordings. Workflow: Invite dev → edit app → merge changes. Metrics: 10+ users.", "Unity Implementation": "Photon Fusion; Agora voice; Unity Netcode.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Collaboration Service", "Notes": "Edge cases: disconnects; effort: 40 person-days."},
#     {"Feature": "AI-Assisted App Development", "Description": "Smart app tools: Sub-elements include procedural UI generation, code suggestion (e.g., input handlers), NLP prompts ('add menu'), platform-specific optimizations, and ethical filters. Workflow: Input prompt → generate code → refine. Metrics: <3s generation.", "Unity Implementation": "ML-Agents v2; Azure OpenAI via backend.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: biased code; effort: 50 person-days."},
#     {"Feature": "App Testing Suite", "Description": "Test apps across platforms: Sub-elements include emulators (Quest/PC/Web), FPS/memory dashboards, automated input tests, crash reporting, and A/B testing. Workflow: Select platform → run test → export report. Metrics: <1min test.", "Unity Implementation": "Unity Device Simulator v3; custom profilers.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: platform bugs; effort: 30 person-days."},
#     {"Feature": "App Accessibility", "Description": "Inclusive apps: Sub-elements include color-blind modes, subtitle UI, scalable text, voice commands, haptic alternatives, and one-handed controls. Workflow: Enable setting → test app → save. Metrics: WCAG 2.2 AA.", "Unity Implementation": "Unity UI Toolkit; accessibility plugins.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "None", "Notes": "Test with diverse users; effort: 25 person-days."},
#     {"Feature": "Custom Scripting for Apps", "Description": "C# for app logic: Sub-elements include VR code editor, app-specific templates (e.g., state triggers), breakpoints in VR, Git integration, and sandboxed execution. Workflow: Write code → debug → deploy. Metrics: <1s compile.", "Unity Implementation": "Visual Studio Code; VR UI overlays.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Visual Scripting for Apps", "Notes": "Edge cases: runtime errors; effort: 30 person-days."},
#     {"Feature": "App Audio", "Description": "3D audio for apps: Sub-elements include spatial occlusion, reverb zones, MIDI for app soundtracks, volume curves, binaural support, and platform validation. Workflow: Place audio → configure → test. Metrics: <10ms latency.", "Unity Implementation": "Unity Audio Spatializer; FMOD v2.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: echo; effort: 20 person-days."},
#     {"Feature": "App VFX", "Description": "App effects: Sub-elements include particle systems for UI, shader effects (glow), wind/gravity for interactions, collision visuals, and preset library. Workflow: Create effect → adjust → save. Metrics: <5% GPU.", "Unity Implementation": "VFX Graph v14; Shader Graph.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "App Rendering", "Notes": "Edge cases: overdraw; effort: 25 person-days."},
#     {"Feature": "Custom Asset Library", "Description": "App asset management: Sub-elements include async uploads, AI-suggested tags, thumbnail generation, sharing permissions (public/team), versioning, moderation tools, and drag-drop to app projects. Workflow: Upload asset → tag → use in app. Metrics: <2s thumbnail.", "Unity Implementation": "Custom UI; AssetDatabase caching.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service, Auth Service", "Notes": "Edge cases: duplicate assets; effort: 40 person-days."},
#     {"Feature": "App Localization", "Description": "Multi-language apps: Sub-elements include metadata translation, UI string keys, locale detection, RTL support, .po exports, and translation memory. Workflow: Set locale → update app → export. Metrics: <1s switch.", "Unity Implementation": "Unity Localization v1.4; VR UI.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "20 languages; edge cases: non-Latin; effort: 20 person-days."},
#     {"Feature": "App Performance Tools", "Description": "App profiling: Sub-elements include FPS graphs, memory trackers, draw call analyzers, platform-specific metrics (ASW), and optimization suggestions. Workflow: Open profiler → analyze → optimize. Metrics: <1% overhead.", "Unity Implementation": "Unity Profiler; VR dashboards.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "App Testing Suite", "Notes": "Edge cases: VR stutter; effort: 25 person-days."},
#     {"Feature": "App Feedback System", "Description": "App feedback: Sub-elements include survey popups, error auto-reporting, feature voting, session replay, and analytics integration. Workflow: Report issue → track resolution. Metrics: <10s submission.", "Unity Implementation": "Unity Analytics; custom modals.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "GDPR-compliant; effort: 15 person-days."},
#     {"Feature": "Cross-Platform SDK Integrator", "Description": "SDK management: Sub-elements include Oculus/SteamVR/OpenXR config UI, version checking, dependency resolution, platform validation, and error logs. Workflow: Select SDK → configure → test. Metrics: <1min setup.", "Unity Implementation": "Unity XR Management; custom UI.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Multi-Platform Build Tools", "Notes": "New; edge cases: SDK conflicts; effort: 25 person-days."},
#     {"Feature": "App State Management", "Description": "App state handling: Sub-elements include save/load UI, cross-device sync, state serialization (JSON), rollback options, and error recovery. Workflow: Save state → sync → restore. Metrics: <1s save.", "Unity Implementation": "Custom state scripts; Unity Cloud.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scene Service", "Notes": "New; edge cases: partial saves; effort: 20 person-days."},
#     {"Feature": "AR App Support", "Description": "Mobile AR apps: Sub-elements include ARKit/ARCore integration, plane detection, anchor placement, AR preview in VR, and export to AR formats. Workflow: Scan environment → build app → test. Metrics: <1s detection.", "Unity Implementation": "AR Foundation v5; XR Management.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "New; edge cases: lighting issues; effort: 30 person-days."},
#     {"Feature": "App Marketplace", "Description": "App asset store: Sub-elements include upload moderation, ratings/reviews, search by tags, payment integration, and sales analytics. Workflow: Submit asset → review → publish. Metrics: <5s search.", "Unity Implementation": "Custom UI; Unity Asset Store API.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "New; edge cases: fraud; effort: 40 person-days."}
# ]

# # Frontend Web Features (focused on VR app development)
# frontend_web = [
#     {"Feature": "WebXR App Preview", "Description": "Preview VR apps in browser: Sub-elements include WebXR session UI, immersive toggles (VR/2D), device permissions, 2D fallback (orbit controls), mobile AR previews, and platform validation. Workflow: Load app → enter WebXR → test. Metrics: <2s session.", "Implementation": "WebXR API 1.1; Unity WebGL v2023; A-Frame.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Test Chrome/Safari; edge cases: permissions; effort: 25 person-days."},
#     {"Feature": "Browser Input Config", "Description": "App input testing: Sub-elements include touch gestures (pinch-zoom), keyboard (WASD), mouse orbit, gamepad support, sensitivity sliders, and multi-platform mappings. Workflow: Configure input → test app → save. Metrics: <10ms latency.", "Implementation": "Unity Input System; WebGL Pointer Lock; Hammer.js.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Edge cases: iOS touch; effort: 20 person-days."},
#     {"Feature": "WebGL App Rendering", "Description": "Render apps in browser: Sub-elements include dynamic resolution, texture mipmapping, shader fallbacks, loading bars, context loss recovery, and platform-specific optimizations. Workflow: Load app → adjust quality → test. Metrics: 60 FPS.", "Implementation": "Unity URP v14; WebGL 2.0 scripts.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "App Rendering", "Notes": "Edge cases: context loss; effort: 25 person-days."},
#     {"Feature": "Web App Asset Support", "Description": "App assets for WebXR: Sub-elements include GLTF validation, poly reduction (50k), streaming loads, format converters, cache management, and CORS handlers. Workflow: Upload asset → validate → use in app. Metrics: <3s load.", "Implementation": "GLTFast v5; Web Workers.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Edge cases: CORS; effort: 20 person-days."},
#     {"Feature": "Web App PWA", "Description": "Installable app tester: Sub-elements include manifest configs, offline caching, service worker updates, push notifications, and install prompts. Workflow: Visit site → install → test offline. Metrics: <5MB cache.", "Implementation": "Workbox; Unity WebGL PWA.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: storage limits; effort: 15 person-days."},
#     {"Feature": "Web App Collaboration", "Description": "Co-develop apps in browser: Sub-elements include cursor sharing, chat overlays, version locking, invite links, tab sync, and conflict UI. Workflow: Invite → edit app → merge. Metrics: <50ms sync.", "Implementation": "WebSockets; Firebase; Yjs.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Multi-User App Collaboration", "Notes": "Edge cases: tab crashes; effort: 30 person-days."},
#     {"Feature": "Web App Analytics", "Description": "Track app usage: Sub-elements include event funnels, heatmaps, A/B test APIs, privacy opt-outs, CSV exports, and dashboards. Workflow: Track session → generate report. Metrics: <1s log.", "Implementation": "Google Analytics 4; custom WebGL events.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Analytics Service", "Notes": "Edge cases: ad blockers; effort: 15 person-days."},
#     {"Feature": "Web App Accessibility", "Description": "Inclusive app testing: Sub-elements include ARIA labels, keyboard navigation, screen reader support, contrast checkers, and reduced motion. Workflow: Enable accessibility → test app. Metrics: WCAG 2.2 AA.", "Implementation": "ARIA; WebGL polyfills.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "App Accessibility", "Notes": "Test with NVDA; effort: 20 person-days."},
#     {"Feature": "Web App Notifications", "Description": "Browser alerts for apps: Sub-elements include permission dialogs, customizable channels, vibration, badge updates, and offline queuing. Workflow: Subscribe → receive alert. Metrics: <2s delivery.", "Implementation": "Web Push API; Service Workers.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Notification Service", "Notes": "Edge cases: denied permissions; effort: 10 person-days."},
#     {"Feature": "Web App Audio", "Description": "App audio in browser: Sub-elements include WebAudio spatial nodes, context resumption, format fallbacks, volume normalization, and debug tools. Workflow: Load audio → test in app. Metrics: <10ms latency.", "Implementation": "WebAudio API; Unity audio.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "App Audio", "Notes": "Edge cases: muted tabs; effort: 15 person-days."},
#     {"Feature": "Custom Asset Library (Web)", "Description": "Web dashboard for app assets: Sub-elements include drag-drop uploads, WebGL previews, metadata editor, sharing links, collaborative editing, and Unity exports. Workflow: Upload → tag → share. Metrics: <3s preview.", "Implementation": "React 18; Tailwind; WebGL shaders.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library (Backend)", "Notes": "Edge cases: large uploads; effort: 30 person-days."},
#     {"Feature": "Web App Localization", "Description": "Multi-language app testing: Sub-elements include locale detection, string loading, RTL layouts, currency/date formatting, and translation APIs. Workflow: Detect locale → load app UI. Metrics: <1s switch.", "Implementation": "react-i18next; WebGL text.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "App Localization", "Notes": "15 languages; effort: 15 person-days."},
#     {"Feature": "Web App Performance", "Description": "Browser app profiling: Sub-elements include Lighthouse audits, network throttles, asset load timelines, error consoles, and optimization reports. Workflow: Run audit → optimize app. Metrics: <1% overhead.", "Implementation": "Chrome DevTools; WebGL profilers.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "App Performance Tools", "Notes": "Edge cases: slow 3G; effort: 20 person-days."},
#     {"Feature": "Web AR App Support", "Description": "AR app previews: Sub-elements include WebXR AR mode, plane detection, anchor UI, camera passthrough, and AR exports. Workflow: Scan room → test app. Metrics: <1s detection.", "Implementation": "WebXR AR; 8th Wall.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "AR App Support", "Notes": "Edge cases: lighting; effort: 25 person-days."}
# ]

# # Backend Features (focused on VR app development)
# backend_features = [
#     {"Component": "API Gateway", "Description": "App API routing: Sub-elements include rate limiting (100 req/s/user), versioning (v1/v2), audit logging, CORS for Unity/Web, health checks, and DDoS throttling. Endpoints: GET /health, POST /auth/token. Metrics: <100ms latency.", "Technologies": "AWS API Gateway; Kong; OAuth2/JWT.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "99.99% uptime; edge cases: traffic spikes; effort: 30 person-days."},
#     {"Component": "Auth Service", "Description": "App user management: Sub-elements include passwordless login, 2FA (TOTP), RBAC (developer/publisher), session revocation, audit trails, and SSO. Endpoints: POST /login, GET /users/me. Metrics: <500ms auth.", "Technologies": "Firebase Auth; Keycloak; Postgres.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "GDPR export; edge cases: token theft; effort: 25 person-days."},
#     {"Component": "Asset Service", "Description": "App asset handling: Sub-elements include virus scanning, metadata indexing, thumbnail generation, deduplication, purge APIs, and CDN caching. Endpoints: POST /assets/upload, GET /assets/{id}. Metrics: <5s for 100MB.", "Technologies": "S3; DynamoDB; FFmpeg; Cloudflare.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "1GB+ files; edge cases: corrupt uploads; effort: 35 person-days."},
#     {"Component": "App State Service", "Description": "App state management: Sub-elements include JSON schema validation, version diffs, autosave (10s), rollback endpoints, USD export, and conflict detection. Endpoints: POST /states, GET /states/{id}/versions. Metrics: <1s save.", "Technologies": "MongoDB 6; libgit2; Redis.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "10MB+ states; edge cases: conflicts; effort: 30 person-days."},
#     {"Component": "Collaboration Service", "Description": "App co-development: Sub-elements include OT for code sync, presence indicators, chat persistence, bandwidth throttling, offline queues, and analytics. Endpoints: WS /collab/join. Metrics: <50ms sync.", "Technologies": "Socket.io; Yjs; Redis Pub/Sub.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "App State Service", "Notes": "50 users; edge cases: disconnects; effort: 40 person-days."},
#     {"Component": "Analytics Service", "Description": "App usage tracking: Sub-elements include funnels (code-to-publish), cohort analysis, crash reports, dashboards, CSV exports, and GDPR opt-outs. Endpoints: POST /events, GET /reports. Metrics: <1s log.", "Technologies": "Elasticsearch 8; Kafka; Kibana.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: large datasets; effort: 25 person-days."},
#     {"Component": "Publishing Service", "Description": "App deployment: Sub-elements include CI/CD pipelines (APK/WebGL/EXE), build queues, version tagging, rollback APIs, and platform analytics. Endpoints: POST /builds, GET /builds/{id}. Metrics: <5min build.", "Technologies": "Jenkins; Unity Cloud Build; Docker.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: build failures; effort: 30 person-days."},
#     {"Component": "Database Layer", "Description": "App data: Sub-elements include sharding, backups (daily), index optimization, Redis caching (1hr TTL), migrations, and HA configs. Schema: users, states, assets. Metrics: <10ms query.", "Technologies": "Postgres 15; MongoDB 6; Redis 7.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "99.9% uptime; edge cases: corruption; effort: 35 person-days."},
#     {"Component": "AI Integration", "Description": "AI for app dev: Sub-elements include prompt templating, rate limiting, asset/code caching, model selection (GPT-4o), billing trackers, and error handling. Endpoints: POST /ai/generate. Metrics: <3s response.", "Technologies": "Azure OpenAI; Hugging Face; SQS.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: API downtime; effort: 40 person-days."},
#     {"Component": "Security & Compliance", "Description": "App security: Sub-elements include AES-256 encryption, OWASP scans, GDPR/CCPA/HIPAA reports, WAF, audit logs, and anonymization. Endpoints: GET /compliance/report. Metrics: 0 vulnerabilities.", "Technologies": "HTTPS/TLS 1.3; OWASP; CloudTrail.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Quarterly pentests; edge cases: DDoS; effort: 30 person-days."},
#     {"Component": "Scalability Layer", "Description": "App scaling: Sub-elements include pod scaling, load balancer configs, stress tests (Locust), metrics (CPU/RAM), cost dashboards, and failover scripts. Metrics: 10k users.", "Technologies": "Kubernetes 1.27; Prometheus; Grafana.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: spikes; effort: 35 person-days."},
#     {"Component": "Notification Service", "Description": "App alerts: Sub-elements include template rendering, subscription UI, retry logic, analytics, SMS (Twilio), and spam filters. Endpoints: POST /notifications. Metrics: <2s delivery.", "Technologies": "Firebase Cloud Messaging; AWS SES.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: undeliverable; effort: 15 person-days."},
#     {"Component": "Custom Asset Library (Backend)", "Description": "App asset storage: Sub-elements include ACLs, versioning, Elasticsearch indexing, auto-cleanup, bulk APIs, moderation, and audit logs. Endpoints: POST /library/upload, GET /library/search. Metrics: <500ms search.", "Technologies": "S3; DynamoDB; Elasticsearch 8.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service, Auth Service", "Notes": "100TB scale; edge cases: conflicts; effort: 40 person-days."},
#     {"Component": "App State Recovery", "Description": "App crash recovery: Sub-elements include autosave (10s), diff-based restore, cross-device sync, error logs, and recovery UI. Endpoints: POST /states/restore. Metrics: <5s restore.", "Technologies": "MongoDB; Redis; S3.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "App State Service", "Notes": "Edge cases: partial saves; effort: 20 person-days."}
# ]

# # Product Launch Plan (focused on VR app development)
# launch_plan = [
#     {"Task": "Define App Developer Audience", "Description": "Research VR app devs: Sub-elements include surveys (indie vs. enterprise needs), personas (game dev, training), competitor analysis (Unity, Godot), platform segmentation (Quest/WebXR), and use case validation (games, sims). Workflow: Survey → define segments. Metrics: 500+ responses.", "Details": "SurveyMonkey; X polls; VR dev forums.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Nov 2025", "Dependencies": "None", "Notes": "Edge cases: niche apps; effort: 20 person-days."},
#     {"Task": "Beta Testing for Apps", "Description": "Test VR apps: Sub-elements include alpha (20 devs), public beta (200 devs), feedback forms, bug triage (Jira), A/B feature tests, and platform testing (Quest/PC/Web). Workflow: Deploy beta → iterate. Metrics: <5% crash rate.", "Details": "TestFlight; Steam Beta; App Lab.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "App Testing Suite", "Notes": "Edge cases: low-end devices; effort: 40 person-days."},
#     {"Task": "Monetization for App Devs", "Description": "Revenue models: Sub-elements include freemium ($20/month pro), in-app asset purchases, enterprise licenses ($5k/year), affiliate links, and A/B pricing tests. Workflow: Set pricing → test → launch. Metrics: 10% conversion.", "Details": "Stripe; PayPal; analytics.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Auth Service", "Notes": "Edge cases: refunds; effort: 30 person-days."},
#     {"Task": "App Compliance", "Description": "Legal standards: Sub-elements include GDPR mapping, WCAG 2.2 audits, HIPAA for sims, CCPA compliance, privacy policies, and certifications. Workflow: Audit → remediate → certify. Metrics: 0 issues.", "Details": "Legal reviews; Axe tool.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Security & Compliance", "Notes": "Annual audits; effort: 25 person-days."},
#     {"Task": "App Marketing", "Description": "Promote app tool: Sub-elements include X/Reddit teasers, YouTube demos, VR streamer partnerships, email campaigns, virtual launch events, and $10k ad budget. Workflow: Plan → execute → track. Metrics: 1M impressions.", "Details": "Hootsuite; Google Ads; Siggraph.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: negative feedback; effort: 35 person-days."},
#     {"Task": "Post-Launch App Support", "Description": "Maintain app tool: Sub-elements include monthly patches, Discord forums, guides (video/text), churn analysis (5% target), feature voting, and roadmaps. Workflow: Monitor → patch → communicate. Metrics: <24hr response.", "Details": "GitHub Issues; Zendesk.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "Edge cases: breaking updates; effort: 30 person-days."},
#     {"Task": "App Scalability Testing", "Description": "Load test app tool: Sub-elements include spike tests (10k users), DB stress tests, failover drills, cost projections ($5k/month), and optimization scripts. Workflow: Simulate → optimize. Metrics: 99.99% uptime.", "Details": "Locust; AWS; New Relic.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scalability Layer", "Notes": "Edge cases: outages; effort: 25 person-days."}
# ]

# # Completed Features (adjusted for app focus)
# completed_features = [
#     {"Sno": 1, "Feature": "App Hierarchy", "Description": "Manage app objects: Sub-elements include tree view (expand/collapse), search by name/type, drag-reorder, multi-select delete, and linking for app logic. Workflow: Open hierarchy → reorder → save. Metrics: <100ms refresh.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Integrate with Custom Asset Library; effort: 10 person-days."},
#     {"Sno": 2, "Feature": "App Scene", "Description": "App logic container: Sub-elements include scene loading, multi-scene support, layer management, prefab instantiation, and save/load UI. Workflow: Create scene → add logic → save. Metrics: <2s load.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Add multi-user sync; effort: 15 person-days."},
#     {"Sno": 3, "Feature": "Move App Object", "Description": "Move app objects: Sub-elements include gizmo controls, numeric input, snap-to-grid, multi-object move, and undo. Workflow: Select → move → confirm. Metrics: <10ms response.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "App Scene", "Notes": "Add VR gestures; effort: 5 person-days."},
#     # ... (include all 44; abbreviated here)
#     {"Sno": 44, "Feature": "Reset App Object", "Description": "Reset app object transforms: Sub-elements include one-click zeroing, batch reset, undo, confirmation dialogs, and gizmo visuals. Workflow: Select → reset → confirm. Metrics: <100ms action.", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "App Scene", "Notes": "Add VR button; effort: 5 person-days."}
# ]

# # Sub-Features Breakdown (focused on app development)
# sub_features = [
#     {"Parent Feature": "App Logic Editor", "Sub-Feature": "State Machine UI", "Description": "Drag-drop state creation; supports transitions, triggers, and platform-specific states. Metrics: <1s transition.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Visual Scripting for Apps", "Notes": "Edge cases: state conflicts; effort: 10 person-days."},
#     {"Parent Feature": "App Logic Editor", "Sub-Feature": "Event Triggers", "Description": "UI for onInput/onStateChange; supports custom events, VR previews. Metrics: <10ms trigger.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Edge cases: event loops; effort: 8 person-days."},
#     {"Parent Feature": "Multi-Platform Build Tools", "Sub-Feature": "SDK Injection", "Description": "Auto-inject Oculus/SteamVR SDKs; includes version checks, dependency resolution. Metrics: <1min setup.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Cross-Platform SDK Integrator", "Notes": "Edge cases: version mismatches; effort: 10 person-days."},
#     {"Parent Feature": "Multi-Platform Build Tools", "Sub-Feature": "Build Validation", "Description": "Validate app for platform specs (e.g., Quest poly limits); includes error logs. Metrics: <30s validation.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Edge cases: invalid configs; effort: 8 person-days."},
#     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Asset Tagging", "Description": "AI-suggested tags for app assets; supports hierarchies, search facets. Metrics: <500ms search.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "AI Integration", "Notes": "Edge cases: duplicates; effort: 10 person-days."},
#     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Versioning", "Description": "Track asset versions; supports diffs, rollback, comments. Metrics: <1s diff load.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "App State Service", "Notes": "Edge cases: conflicts; effort: 8 person-days."},
#     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Sharing Permissions", "Description": "ACLs for app assets; includes invite links, role-based access. Metrics: <200ms check.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Auth Service", "Notes": "Edge cases: revoked access; effort: 10 person-days."},
#     {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Moderation Tools", "Description": "Flag/review app assets; includes admin UI, AI flagging. Metrics: <1s processing.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: false positives; effort: 7 person-days."},
#     {"Parent Feature": "WebXR App Preview", "Sub-Feature": "Session Initiation", "Description": "Enter WebXR for app testing; includes device detection, error UI. Metrics: <2s start.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: unsupported browsers; effort: 5 person-days."},
#     {"Parent Feature": "WebXR App Preview", "Sub-Feature": "Immersive Toggles", "Description": "Switch VR/2D for app testing; includes UI animations. Metrics: <1s toggle.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Browser Input Config", "Notes": "Edge cases: mode conflicts; effort: 5 person-days."}
#     # Add ~100 more sub-features in full implementation
# ]

# # Create sheets and populate data
# sheets_data = {
#     "Frontend Unity": frontend_unity,
#     "Frontend Web": frontend_web,
#     "Backend": backend_features,
#     "Product Launch Plan": launch_plan,
#     "Completed Features": completed_features,
#     "Sub-Features Breakdown": sub_features
# }

# # Process each sheet
# for sheet_name, data in sheets_data.items():
#     ws = workbook.create_sheet(title=sheet_name)
#     headers = list(data[0].keys())
    
#     # Write headers
#     for col_idx, header in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col_idx)
#         cell.value = header
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
#     # Sort data by Priority
#     sorted_data = sorted(data, key=lambda x: priority_order.get(x["Priority"], 3))
    
#     # Write data
#     for row_idx, row_data in enumerate(sorted_data, start=2):
#         for col_idx, key in enumerate(headers, start=1):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             cell.value = row_data[key]
#             cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
#             # Conditional formatting for overdue tasks
#             timeline_str = row_data.get("Timeline", "")
#             try:
#                 timeline_date = datetime.strptime(timeline_str, "%b %Y")
#                 if timeline_date < current_date and row_data.get("Status", "") != "Completed":
#                     cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
#             except:
#                 pass
    
#     # Apply filters
#     ws.auto_filter.ref = ws.dimensions
    
#     # Auto-adjust column widths
#     for col_idx in range(1, len(headers) + 1):
#         max_length = 0
#         column_letter = get_column_letter(col_idx)
#         for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
#             for cell in row:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#         adjusted_width = min((max_length + 2) * 1.2, 70)
#         ws.column_dimensions[column_letter].width = adjusted_width

# # Save the workbook
# workbook.save("VR_App_Authoring_Tool_Roadmap.xlsx")
# print("Excel file 'VR_App_Authoring_Tool_Roadmap.xlsx' has been created successfully.")
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a new workbook
workbook = openpyxl.Workbook()
workbook.remove(workbook.active)

# Define priority order for sorting
priority_order = {"High": 1, "Medium": 2, "Low": 3}

# Current date for conditional formatting
current_date = datetime(2025, 8, 26)

# Frontend Unity Features (focused on VR app development)
frontend_unity = [
    {"Feature": "App Logic Editor", "Description": "Visual editor for VR app logic: Sub-elements include state machine UI (drag-drop states), event triggers (onInput/onEvent), cross-platform input mappings (Oculus/SteamVR), scriptable logic templates, and error checking for state conflicts. Workflow: Create state → link events → test on platform. Metrics: <1s state transition.", "Unity Implementation": "XR Interaction Toolkit v2.5; Unity Visual Scripting for states; OpenXR input mappings.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library, Scene Service", "Notes": "Core for app devs; test Quest 3/SteamVR; edge cases: input conflicts; effort: 30 person-days."},
    {"Feature": "Multi-Platform Build Tools", "Description": "Build VR apps for Quest/PC/WebXR: Sub-elements include build config UI (APK/EXE/WebGL), automated SDK injection (Oculus/SteamVR), platform-specific optimizations (e.g., Quest LODs), build validation, and error logs. Workflow: Select platform → configure → build. Metrics: <5min build.", "Unity Implementation": "Unity Build Pipeline; custom SDK scripts; Unity Cloud Build.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Edge cases: SDK version mismatches; effort: 35 person-days."},
    {"Feature": "Interactable System", "Description": "App interactions: Sub-elements include haptic feedback (variable intensity), multi-object grab for app controls, platform-specific input bindings, event triggers for app logic, and voice input support. Workflow: Assign interaction → test on platform → save. Metrics: <10ms haptic latency.", "Unity Implementation": "XRGrabInteractable; OpenXR haptics; custom UnityEvent system.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Scene Service", "Notes": "Test on Quest/Vive; edge cases: multi-platform input; effort: 25 person-days."},
    {"Feature": "Animation for Apps", "Description": "App animations: Sub-elements include keyframe UI for UI transitions, skeletal rigging for characters, blend shapes for app states, VR preview for animations, and export to platform formats (GLTF/USD). Workflow: Create animation → test → export. Metrics: <1s export.", "Unity Implementation": "Unity Timeline v1.8; Mixamo; custom VR UI.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: platform-specific playback; effort: 30 person-days."},
    {"Feature": "Physics for Apps", "Description": "App physics: Sub-elements include rigidbody for interactive controls, soft-body for UI effects, collision events for app logic, raycasting for selections, and platform-optimized physics. Workflow: Add physics → configure → test. Metrics: <5% CPU.", "Unity Implementation": "Unity Physics (DOTS); NVIDIA PhysX 5.0.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Scene Service", "Notes": "Optimize for Quest; edge cases: high-velocity; effort: 30 person-days."},
    {"Feature": "Visual Scripting for Apps", "Description": "No-code app logic: Sub-elements include node-based UI for app flows, VR event nodes (onInput/onStateChange), debugging tools, templates for app patterns (menus, inventory), and exportable scripts. Workflow: Create nodes → test → deploy. Metrics: <1s execution.", "Unity Implementation": "Unity Visual Scripting v1.8; custom VR nodes.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: cyclic graphs; effort: 20 person-days."},
    {"Feature": "App Rendering", "Description": "Cross-platform rendering: Sub-elements include URP/HDRP toggles, platform-specific shaders (Quest/PC), light baking for apps, volumetric effects for UI, and adaptive quality. Workflow: Configure render → test platform → optimize. Metrics: 90 FPS.", "Unity Implementation": "URP v14; Bakery plugin; custom VR shaders.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: mobile performance; effort: 25 person-days."},
    {"Feature": "Input Configurator", "Description": "Platform input mappings: Sub-elements include gesture recognition (pinch/swipe), controller bindings (Oculus/SteamVR), keyboard/gamepad support, input remapping UI, and calibration tools. Workflow: Map input → test → save. Metrics: >95% accuracy.", "Unity Implementation": "OpenXR 1.9; Oculus Integration v50.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Interactable System", "Notes": "Edge cases: input conflicts; effort: 20 person-days."},
    {"Feature": "Multi-User App Collaboration", "Description": "Co-develop apps: Sub-elements include avatar systems, spatial voice chat, real-time code sync (<50ms), conflict resolution UI, and session recordings. Workflow: Invite dev → edit app → merge changes. Metrics: 10+ users.", "Unity Implementation": "Photon Fusion; Agora voice; Unity Netcode.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Collaboration Service", "Notes": "Edge cases: disconnects; effort: 40 person-days."},
    {"Feature": "AI-Assisted App Development", "Description": "Smart app tools: Sub-elements include procedural UI generation, code suggestion (e.g., input handlers), NLP prompts ('add menu'), platform-specific optimizations, and ethical filters. Workflow: Input prompt → generate code → refine. Metrics: <3s generation.", "Unity Implementation": "ML-Agents v2; Azure OpenAI via backend.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: biased code; effort: 50 person-days."},
    {"Feature": "App Testing Suite", "Description": "Test apps across platforms: Sub-elements include emulators (Quest/PC/Web), FPS/memory dashboards, automated input tests, crash reporting, and A/B testing. Workflow: Select platform → run test → export report. Metrics: <1min test.", "Unity Implementation": "Unity Device Simulator v3; custom profilers.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: platform bugs; effort: 30 person-days."},
    {"Feature": "App Accessibility", "Description": "Inclusive apps: Sub-elements include color-blind modes, subtitle UI, scalable text, voice commands, haptic alternatives, and one-handed controls. Workflow: Enable setting → test app → save. Metrics: WCAG 2.2 AA.", "Unity Implementation": "Unity UI Toolkit; accessibility plugins.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "None", "Notes": "Test with diverse users; effort: 25 person-days."},
    {"Feature": "Custom Scripting for Apps", "Description": "C# for app logic: Sub-elements include VR code editor, app-specific templates (e.g., state triggers), breakpoints in VR, Git integration, and sandboxed execution. Workflow: Write code → debug → deploy. Metrics: <1s compile.", "Unity Implementation": "Visual Studio Code; VR UI overlays.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Visual Scripting for Apps", "Notes": "Edge cases: runtime errors; effort: 30 person-days."},
    {"Feature": "App Audio", "Description": "3D audio for apps: Sub-elements include spatial occlusion, reverb zones, MIDI for app soundtracks, volume curves, binaural support, and platform validation. Workflow: Place audio → configure → test. Metrics: <10ms latency.", "Unity Implementation": "Unity Audio Spatializer; FMOD v2.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: echo; effort: 20 person-days."},
    {"Feature": "App VFX", "Description": "App effects: Sub-elements include particle systems for UI, shader effects (glow), wind/gravity for interactions, collision visuals, and preset library. Workflow: Create effect → adjust → save. Metrics: <5% GPU.", "Unity Implementation": "VFX Graph v14; Shader Graph.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "App Rendering", "Notes": "Edge cases: overdraw; effort: 25 person-days."},
    {"Feature": "Custom Asset Library", "Description": "App asset management: Sub-elements include async uploads, AI-suggested tags, thumbnail generation, sharing permissions (public/team), versioning, moderation tools, and drag-drop to app projects. Workflow: Upload asset → tag → use in app. Metrics: <2s thumbnail.", "Unity Implementation": "Custom UI; AssetDatabase caching.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service, Auth Service", "Notes": "Edge cases: duplicate assets; effort: 40 person-days."},
    {"Feature": "App Localization", "Description": "Multi-language apps: Sub-elements include metadata translation, UI string keys, locale detection, RTL support, .po exports, and translation memory. Workflow: Set locale → update app → export. Metrics: <1s switch.", "Unity Implementation": "Unity Localization v1.4; VR UI.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "20 languages; edge cases: non-Latin; effort: 20 person-days."},
    {"Feature": "App Performance Tools", "Description": "App profiling: Sub-elements include FPS graphs, memory trackers, draw call analyzers, platform-specific metrics (ASW), and optimization suggestions. Workflow: Open profiler → analyze → optimize. Metrics: <1% overhead.", "Unity Implementation": "Unity Profiler; VR dashboards.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "App Testing Suite", "Notes": "Edge cases: VR stutter; effort: 25 person-days."},
    {"Feature": "App Feedback System", "Description": "App feedback: Sub-elements include survey popups, error auto-reporting, feature voting, session replay, and analytics integration. Workflow: Report issue → track resolution. Metrics: <10s submission.", "Unity Implementation": "Unity Analytics; custom modals.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "GDPR-compliant; effort: 15 person-days."},
    {"Feature": "Cross-Platform SDK Integrator", "Description": "SDK management: Sub-elements include Oculus/SteamVR/OpenXR config UI, version checking, dependency resolution, platform validation, and error logs. Workflow: Select SDK → configure → test. Metrics: <1min setup.", "Unity Implementation": "Unity XR Management; custom UI.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Multi-Platform Build Tools", "Notes": "New; edge cases: SDK conflicts; effort: 25 person-days."},
    {"Feature": "App State Management", "Description": "App state handling: Sub-elements include save/load UI, cross-device sync, state serialization (JSON), rollback options, and error recovery. Workflow: Save state → sync → restore. Metrics: <1s save.", "Unity Implementation": "Custom state scripts; Unity Cloud.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scene Service", "Notes": "New; edge cases: partial saves; effort: 20 person-days."},
    {"Feature": "AR App Support", "Description": "Mobile AR apps: Sub-elements include ARKit/ARCore integration, plane detection, anchor placement, AR preview in VR, and export to AR formats. Workflow: Scan environment → build app → test. Metrics: <1s detection.", "Unity Implementation": "AR Foundation v5; XR Management.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "New; edge cases: lighting issues; effort: 30 person-days."},
    {"Feature": "App Marketplace", "Description": "App asset store: Sub-elements include upload moderation, ratings/reviews, search by tags, payment integration, and sales analytics. Workflow: Submit asset → review → publish. Metrics: <5s search.", "Unity Implementation": "Custom UI; Unity Asset Store API.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "Custom Asset Library", "Notes": "New; edge cases: fraud; effort: 40 person-days."}
]

# Frontend Web Features (focused on VR app development)
frontend_web = [
    {"Feature": "WebXR App Preview", "Description": "Preview VR apps in browser: Sub-elements include WebXR session UI, immersive toggles (VR/2D), device permissions, 2D fallback (orbit controls), mobile AR previews, and platform validation. Workflow: Load app → enter WebXR → test. Metrics: <2s session.", "Implementation": "WebXR API 1.1; Unity WebGL v2023; A-Frame.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Test Chrome/Safari; edge cases: permissions; effort: 25 person-days."},
    {"Feature": "Browser Input Config", "Description": "App input testing: Sub-elements include touch gestures (pinch-zoom), keyboard (WASD), mouse orbit, gamepad support, sensitivity sliders, and multi-platform mappings. Workflow: Configure input → test app → save. Metrics: <10ms latency.", "Implementation": "Unity Input System; WebGL Pointer Lock; Hammer.js.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Edge cases: iOS touch; effort: 20 person-days."},
    {"Feature": "WebGL App Rendering", "Description": "Render apps in browser: Sub-elements include dynamic resolution, texture mipmapping, shader fallbacks, loading bars, context loss recovery, and platform-specific optimizations. Workflow: Load app → adjust quality → test. Metrics: 60 FPS.", "Implementation": "Unity URP v14; WebGL 2.0 scripts.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "App Rendering", "Notes": "Edge cases: context loss; effort: 25 person-days."},
    {"Feature": "Web App Asset Support", "Description": "App assets for WebXR: Sub-elements include GLTF validation, poly reduction (50k), streaming loads, format converters, cache management, and CORS handlers. Workflow: Upload asset → validate → use in app. Metrics: <3s load.", "Implementation": "GLTFast v5; Web Workers.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service", "Notes": "Edge cases: CORS; effort: 20 person-days."},
    {"Feature": "Web App PWA", "Description": "Installable app tester: Sub-elements include manifest configs, offline caching, service worker updates, push notifications, and install prompts. Workflow: Visit site → install → test offline. Metrics: <5MB cache.", "Implementation": "Workbox; Unity WebGL PWA.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: storage limits; effort: 15 person-days."},
    {"Feature": "Web App Collaboration", "Description": "Co-develop apps in browser: Sub-elements include cursor sharing, chat overlays, version locking, invite links, tab sync, and conflict UI. Workflow: Invite → edit app → merge. Metrics: <50ms sync.", "Implementation": "WebSockets; Firebase; Yjs.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Multi-User App Collaboration", "Notes": "Edge cases: tab crashes; effort: 30 person-days."},
    {"Feature": "Web App Analytics", "Description": "Track app usage: Sub-elements include event funnels, heatmaps, A/B test APIs, privacy opt-outs, CSV exports, and dashboards. Workflow: Track session → generate report. Metrics: <1s log.", "Implementation": "Google Analytics 4; custom WebGL events.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Analytics Service", "Notes": "Edge cases: ad blockers; effort: 15 person-days."},
    {"Feature": "Web App Accessibility", "Description": "Inclusive app testing: Sub-elements include ARIA labels, keyboard navigation, screen reader support, contrast checkers, and reduced motion. Workflow: Enable accessibility → test app. Metrics: WCAG 2.2 AA.", "Implementation": "ARIA; WebGL polyfills.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "App Accessibility", "Notes": "Test with NVDA; effort: 20 person-days."},
    {"Feature": "Web App Notifications", "Description": "Browser alerts for apps: Sub-elements include permission dialogs, customizable channels, vibration, badge updates, and offline queuing. Workflow: Subscribe → receive alert. Metrics: <2s delivery.", "Implementation": "Web Push API; Service Workers.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Notification Service", "Notes": "Edge cases: denied permissions; effort: 10 person-days."},
    {"Feature": "Web App Audio", "Description": "App audio in browser: Sub-elements include WebAudio spatial nodes, context resumption, format fallbacks, volume normalization, and debug tools. Workflow: Load audio → test in app. Metrics: <10ms latency.", "Implementation": "WebAudio API; Unity audio.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "App Audio", "Notes": "Edge cases: muted tabs; effort: 15 person-days."},
    {"Feature": "Custom Asset Library (Web)", "Description": "Web dashboard for app assets: Sub-elements include drag-drop uploads, WebGL previews, metadata editor, sharing links, collaborative editing, and Unity exports. Workflow: Upload → tag → share. Metrics: <3s preview.", "Implementation": "React 18; Tailwind; WebGL shaders.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Custom Asset Library (Backend)", "Notes": "Edge cases: large uploads; effort: 30 person-days."},
    {"Feature": "Web App Localization", "Description": "Multi-language app testing: Sub-elements include locale detection, string loading, RTL layouts, currency/date formatting, and translation APIs. Workflow: Detect locale → load app UI. Metrics: <1s switch.", "Implementation": "react-i18next; WebGL text.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "App Localization", "Notes": "15 languages; effort: 15 person-days."},
    {"Feature": "Web App Performance", "Description": "Browser app profiling: Sub-elements include Lighthouse audits, network throttles, asset load timelines, error consoles, and optimization reports. Workflow: Run audit → optimize app. Metrics: <1% overhead.", "Implementation": "Chrome DevTools; WebGL profilers.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "App Performance Tools", "Notes": "Edge cases: slow 3G; effort: 20 person-days."},
    {"Feature": "Web AR App Support", "Description": "AR app previews: Sub-elements include WebXR AR mode, plane detection, anchor UI, camera passthrough, and AR exports. Workflow: Scan room → test app. Metrics: <1s detection.", "Implementation": "WebXR AR; 8th Wall.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "AR App Support", "Notes": "Edge cases: lighting; effort: 25 person-days."}
]

# Backend Features (focused on VR app development)
backend_features = [
    {"Component": "API Gateway", "Description": "App API routing: Sub-elements include rate limiting (100 req/s/user), versioning (v1/v2), audit logging, CORS for Unity/Web, health checks, and DDoS throttling. Endpoints: GET /health, POST /auth/token. Metrics: <100ms latency.", "Technologies": "AWS API Gateway; Kong; OAuth2/JWT.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "99.99% uptime; edge cases: traffic spikes; effort: 30 person-days."},
    {"Component": "Auth Service", "Description": "App user management: Sub-elements include passwordless login, 2FA (TOTP), RBAC (developer/publisher), session revocation, audit trails, and SSO. Endpoints: POST /login, GET /users/me. Metrics: <500ms auth.", "Technologies": "Firebase Auth; Keycloak; Postgres.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "GDPR export; edge cases: token theft; effort: 25 person-days."},
    {"Component": "Asset Service", "Description": "App asset handling: Sub-elements include virus scanning, metadata indexing, thumbnail generation, deduplication, purge APIs, and CDN caching. Endpoints: POST /assets/upload, GET /assets/{id}. Metrics: <5s for 100MB.", "Technologies": "S3; DynamoDB; FFmpeg; Cloudflare.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "1GB+ files; edge cases: corrupt uploads; effort: 35 person-days."},
    {"Component": "App State Service", "Description": "App state management: Sub-elements include JSON schema validation, version diffs, autosave (10s), rollback endpoints, USD export, and conflict detection. Endpoints: POST /states, GET /states/{id}/versions. Metrics: <1s save.", "Technologies": "MongoDB 6; libgit2; Redis.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "10MB+ states; edge cases: conflicts; effort: 30 person-days."},
    {"Component": "Collaboration Service", "Description": "App co-development: Sub-elements include OT for code sync, presence indicators, chat persistence, bandwidth throttling, offline queues, and analytics. Endpoints: WS /collab/join. Metrics: <50ms sync.", "Technologies": "Socket.io; Yjs; Redis Pub/Sub.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "App State Service", "Notes": "50 users; edge cases: disconnects; effort: 40 person-days."},
    {"Component": "Analytics Service", "Description": "App usage tracking: Sub-elements include funnels (code-to-publish), cohort analysis, crash reports, dashboards, CSV exports, and GDPR opt-outs. Endpoints: POST /events, GET /reports. Metrics: <1s log.", "Technologies": "Elasticsearch 8; Kafka; Kibana.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: large datasets; effort: 25 person-days."},
    {"Component": "Publishing Service", "Description": "App deployment: Sub-elements include CI/CD pipelines (APK/WebGL/EXE), build queues, version tagging, rollback APIs, and platform analytics. Endpoints: POST /builds, GET /builds/{id}. Metrics: <5min build.", "Technologies": "Jenkins; Unity Cloud Build; Docker.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Asset Service", "Notes": "Edge cases: build failures; effort: 30 person-days."},
    {"Component": "Database Layer", "Description": "App data: Sub-elements include sharding, backups (daily), index optimization, Redis caching (1hr TTL), migrations, and HA configs. Schema: users, states, assets. Metrics: <10ms query.", "Technologies": "Postgres 15; MongoDB 6; Redis 7.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "API Gateway", "Notes": "99.9% uptime; edge cases: corruption; effort: 35 person-days."},
    {"Component": "AI Integration", "Description": "AI for app dev: Sub-elements include prompt templating, rate limiting, asset/code caching, model selection (GPT-4o), billing trackers, and error handling. Endpoints: POST /ai/generate. Metrics: <3s response.", "Technologies": "Azure OpenAI; Hugging Face; SQS.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: API downtime; effort: 40 person-days."},
    {"Component": "Security & Compliance", "Description": "App security: Sub-elements include AES-256 encryption, OWASP scans, GDPR/CCPA/HIPAA reports, WAF, audit logs, and anonymization. Endpoints: GET /compliance/report. Metrics: 0 vulnerabilities.", "Technologies": "HTTPS/TLS 1.3; OWASP; CloudTrail.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Quarterly pentests; edge cases: DDoS; effort: 30 person-days."},
    {"Component": "Scalability Layer", "Description": "App scaling: Sub-elements include pod scaling, load balancer configs, stress tests (Locust), metrics (CPU/RAM), cost dashboards, and failover scripts. Metrics: 10k users.", "Technologies": "Kubernetes 1.27; Prometheus; Grafana.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Sep 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: spikes; effort: 35 person-days."},
    {"Component": "Notification Service", "Description": "App alerts: Sub-elements include template rendering, subscription UI, retry logic, analytics, SMS (Twilio), and spam filters. Endpoints: POST /notifications. Metrics: <2s delivery.", "Technologies": "Firebase Cloud Messaging; AWS SES.", "Priority": "Low", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "API Gateway", "Notes": "Edge cases: undeliverable; effort: 15 person-days."},
    {"Component": "Custom Asset Library (Backend)", "Description": "App asset storage: Sub-elements include ACLs, versioning, Elasticsearch indexing, auto-cleanup, bulk APIs, moderation, and audit logs. Endpoints: POST /library/upload, GET /library/search. Metrics: <500ms search.", "Technologies": "S3; DynamoDB; Elasticsearch 8.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Asset Service, Auth Service", "Notes": "100TB scale; edge cases: conflicts; effort: 40 person-days."},
    {"Component": "App State Recovery", "Description": "App crash recovery: Sub-elements include autosave (10s), diff-based restore, cross-device sync, error logs, and recovery UI. Endpoints: POST /states/restore. Metrics: <5s restore.", "Technologies": "MongoDB; Redis; S3.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "App State Service", "Notes": "Edge cases: partial saves; effort: 20 person-days."}
]

# Product Launch Plan (focused on VR app development)
launch_plan = [
    {"Task": "Define App Developer Audience", "Description": "Research VR app devs: Sub-elements include surveys (indie vs. enterprise needs), personas (game dev, training), competitor analysis (Unity, Godot), platform segmentation (Quest/WebXR), and use case validation (games, sims). Workflow: Survey → define segments. Metrics: 500+ responses.", "Details": "SurveyMonkey; X polls; VR dev forums.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Nov 2025", "Dependencies": "None", "Notes": "Edge cases: niche apps; effort: 20 person-days."},
    {"Task": "Beta Testing for Apps", "Description": "Test VR apps: Sub-elements include alpha (20 devs), public beta (200 devs), feedback forms, bug triage (Jira), A/B feature tests, and platform testing (Quest/PC/Web). Workflow: Deploy beta → iterate. Metrics: <5% crash rate.", "Details": "TestFlight; Steam Beta; App Lab.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "App Testing Suite", "Notes": "Edge cases: low-end devices; effort: 40 person-days."},
    {"Task": "Monetization for App Devs", "Description": "Revenue models: Sub-elements include freemium ($20/month pro), in-app asset purchases, enterprise licenses ($5k/year), affiliate links, and A/B pricing tests. Workflow: Set pricing → test → launch. Metrics: 10% conversion.", "Details": "Stripe; PayPal; analytics.", "Priority": "High", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "Auth Service", "Notes": "Edge cases: refunds; effort: 30 person-days."},
    {"Task": "App Compliance", "Description": "Legal standards: Sub-elements include GDPR mapping, WCAG 2.2 audits, HIPAA for sims, CCPA compliance, privacy policies, and certifications. Workflow: Audit → remediate → certify. Metrics: 0 issues.", "Details": "Legal reviews; Axe tool.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Security & Compliance", "Notes": "Annual audits; effort: 25 person-days."},
    {"Task": "App Marketing", "Description": "Promote app tool: Sub-elements include X/Reddit teasers, YouTube demos, VR streamer partnerships, email campaigns, virtual launch events, and $10k ad budget. Workflow: Plan → execute → track. Metrics: 1M impressions.", "Details": "Hootsuite; Google Ads; Siggraph.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Publishing Service", "Notes": "Edge cases: negative feedback; effort: 35 person-days."},
    {"Task": "Post-Launch App Support", "Description": "Maintain app tool: Sub-elements include monthly patches, Discord forums, guides (video/text), churn analysis (5% target), feature voting, and roadmaps. Workflow: Monitor → patch → communicate. Metrics: <24hr response.", "Details": "GitHub Issues; Zendesk.", "Priority": "Medium", "Phase": "Phase 3", "Timeline": "Dec 2026", "Dependencies": "Analytics Service", "Notes": "Edge cases: breaking updates; effort: 30 person-days."},
    {"Task": "App Scalability Testing", "Description": "Load test app tool: Sub-elements include spike tests (10k users), DB stress tests, failover drills, cost projections ($5k/month), and optimization scripts. Workflow: Simulate → optimize. Metrics: 99.99% uptime.", "Details": "Locust; AWS; New Relic.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Jun 2026", "Dependencies": "Scalability Layer", "Notes": "Edge cases: outages; effort: 25 person-days."}
]

# Completed Features (adjusted for app focus)
completed_features = [
    {"Sno": 1, "Feature": "App Hierarchy", "Description": "Manage app objects: Sub-elements include tree view (expand/collapse), search by name/type, drag-reorder, multi-select delete, and linking for app logic. Workflow: Open hierarchy → reorder → save. Metrics: <100ms refresh.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Integrate with Custom Asset Library; effort: 10 person-days."},
    {"Sno": 2, "Feature": "App Scene", "Description": "App logic container: Sub-elements include scene loading, multi-scene support, layer management, prefab instantiation, and save/load UI. Workflow: Create scene → add logic → save. Metrics: <2s load.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "None", "Notes": "Add multi-user sync; effort: 15 person-days."},
    {"Sno": 3, "Feature": "Move App Object", "Description": "Move app objects: Sub-elements include gizmo controls, numeric input, snap-to-grid, multi-object move, and undo. Workflow: Select → move → confirm. Metrics: <10ms response.", "Team": "XR", "Status": "Completed", "Priority": "High", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "App Scene", "Notes": "Add VR gestures; effort: 5 person-days."},
    # ... (include all 44; abbreviated here)
    {"Sno": 44, "Feature": "Reset App Object", "Description": "Reset app object transforms: Sub-elements include one-click zeroing, batch reset, undo, confirmation dialogs, and gizmo visuals. Workflow: Select → reset → confirm. Metrics: <100ms action.", "Team": "XR", "Status": "Completed", "Priority": "Medium", "Phase": "Phase 1", "Timeline": "Aug 2025", "Dependencies": "App Scene", "Notes": "Add VR button; effort: 5 person-days."}
]

# Sub-Features Breakdown (focused on app development)
sub_features = [
    {"Parent Feature": "App Logic Editor", "Sub-Feature": "State Machine UI", "Description": "Drag-drop state creation; supports transitions, triggers, and platform-specific states. Metrics: <1s transition.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Visual Scripting for Apps", "Notes": "Edge cases: state conflicts; effort: 10 person-days."},
    {"Parent Feature": "App Logic Editor", "Sub-Feature": "Event Triggers", "Description": "UI for onInput/onStateChange; supports custom events, VR previews. Metrics: <10ms trigger.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Interactable System", "Notes": "Edge cases: event loops; effort: 8 person-days."},
    {"Parent Feature": "Multi-Platform Build Tools", "Sub-Feature": "SDK Injection", "Description": "Auto-inject Oculus/SteamVR SDKs; includes version checks, dependency resolution. Metrics: <1min setup.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Cross-Platform SDK Integrator", "Notes": "Edge cases: version mismatches; effort: 10 person-days."},
    {"Parent Feature": "Multi-Platform Build Tools", "Sub-Feature": "Build Validation", "Description": "Validate app for platform specs (e.g., Quest poly limits); includes error logs. Metrics: <30s validation.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Publishing Service", "Notes": "Edge cases: invalid configs; effort: 8 person-days."},
    {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Asset Tagging", "Description": "AI-suggested tags for app assets; supports hierarchies, search facets. Metrics: <500ms search.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "AI Integration", "Notes": "Edge cases: duplicates; effort: 10 person-days."},
    {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Versioning", "Description": "Track asset versions; supports diffs, rollback, comments. Metrics: <1s diff load.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "App State Service", "Notes": "Edge cases: conflicts; effort: 8 person-days."},
    {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Sharing Permissions", "Description": "ACLs for app assets; includes invite links, role-based access. Metrics: <200ms check.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Auth Service", "Notes": "Edge cases: revoked access; effort: 10 person-days."},
    {"Parent Feature": "Custom Asset Library", "Sub-Feature": "Moderation Tools", "Description": "Flag/review app assets; includes admin UI, AI flagging. Metrics: <1s processing.", "Priority": "Medium", "Phase": "Phase 2", "Timeline": "Mar 2026", "Dependencies": "AI Integration", "Notes": "Edge cases: false positives; effort: 7 person-days."},
    {"Parent Feature": "WebXR App Preview", "Sub-Feature": "Session Initiation", "Description": "Enter WebXR for app testing; includes device detection, error UI. Metrics: <2s start.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "None", "Notes": "Edge cases: unsupported browsers; effort: 5 person-days."},
    {"Parent Feature": "WebXR App Preview", "Sub-Feature": "Immersive Toggles", "Description": "Switch VR/2D for app testing; includes UI animations. Metrics: <1s toggle.", "Priority": "High", "Phase": "Phase 1", "Timeline": "Dec 2025", "Dependencies": "Browser Input Config", "Notes": "Edge cases: mode conflicts; effort: 5 person-days."}
    # Add ~100 more sub-features in full implementation
]

# Create sheets and populate data
sheets_data = {
    "Frontend Unity": frontend_unity,
    "Frontend Web": frontend_web,
    "Backend": backend_features,
    "Product Launch Plan": launch_plan,
    "Completed Features": completed_features,
    "Sub-Features Breakdown": sub_features
}

# Process each sheet
for sheet_name, data in sheets_data.items():
    ws = workbook.create_sheet(title=sheet_name)
    headers = list(data[0].keys())
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Sort data by Priority
    sorted_data = sorted(data, key=lambda x: priority_order.get(x["Priority"], 3))
    
    # Write data
    for row_idx, row_data in enumerate(sorted_data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data[key]
            cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
            # Conditional formatting for overdue tasks
            timeline_str = row_data.get("Timeline", "")
            try:
                timeline_date = datetime.strptime(timeline_str, "%b %Y")
                if timeline_date < current_date and row_data.get("Status", "") != "Completed":
                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            except:
                pass
    
    # Apply filters
    ws.auto_filter.ref = ws.dimensions
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min((max_length + 2) * 1.2, 70)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
workbook.save("VR_App_Authoring_Tool_Roadmap_new.xlsx")
print("Excel file 'VR_App_Authoring_Tool_Roadmap.xlsx' has been created successfully.")