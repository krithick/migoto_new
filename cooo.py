# import openpyxl
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
# from openpyxl.utils import get_column_letter
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.chart import PieChart, BarChart, Reference
# from datetime import datetime, timedelta

# # Create a new workbook
# wb = openpyxl.Workbook()

# # Remove the default sheet
# wb.remove(wb.active)

# # --- POC-Immerz Sheet ---
# sheet1 = wb.create_sheet("POC-Immerz")

# # Monthly Costs Data
# monthly_data = [
#     {"date": 45778, "usd": 2.63694215646281, "inr": 224.222487741728, "currency": "INR"},
#     {"date": 45809, "usd": 88.7876162314574, "inr": 7578.68895247666, "currency": "INR"},
#     {"date": 45839, "usd": 109.400520267766, "inr": 9376.03483889856, "currency": "INR"},
#     {"date": 45870, "usd": 109.560839140098, "inr": 9512.07205414331, "currency": "INR"},
#     {"date": 45901, "usd": 27.6876505610206, "inr": 2427.86085856949, "currency": "INR"},
# ]

# # Resource Breakdown Data
# resource_data = [
#     {"resource": "novacimmerz-search-aivr-dev (microsoft.search/searchservices)", "cost": 24126.59},
#     {"resource": "novacimmerze-dev-openai-aivr (microsoft.cognitiveservices/accounts)", "cost": 4419.72},
#     {"resource": "novacimmerz-speech-aivr-dev (microsoft.cognitiveservices/accounts)", "cost": 643.39},
#     {"resource": "novacimmerzaivrdev (microsoft.storage/storageaccounts)", "cost": 0.11},
# ]

# # Write Sheet Title
# sheet1["A1"] = "POC-Immerz Azure Costs"
# sheet1["A1"].font = Font(bold=True, size=14)
# sheet1["A1"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# # Write Monthly Costs Headers
# monthly_headers = ["Usage Date", "Cost (USD)", "Cost (INR)", "Currency"]
# for col, header in enumerate(monthly_headers, 1):
#     cell = sheet1.cell(row=3, column=col)
#     cell.value = header
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal="center")
#     cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# # Write Monthly Costs Data
# base_date = datetime(1899, 12, 30)  # Excel date origin
# for row, data in enumerate(monthly_data, 4):
#     sheet1[f"A{row}"] = base_date + timedelta(days=data["date"])
#     sheet1[f"B{row}"] = data["usd"]
#     sheet1[f"C{row}"] = data["inr"]
#     sheet1[f"D{row}"] = data["currency"]

# # Format Monthly Costs
# for row in range(4, 9):
#     sheet1[f"A{row}"].number_format = "yyyy-mm-dd"
#     sheet1[f"B{row}"].number_format = "$#,##0.00"
#     sheet1[f"C{row}"].number_format = "₹#,##0.00"
#     for col in range(1, 5):
#         sheet1.cell(row=row, column=col).alignment = Alignment(horizontal="center")
#         sheet1.cell(row=row, column=col).border = Border(
#             left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
#         )

# # Apply Conditional Formatting to Cost (INR)
# sheet1.conditional_formatting.add(
#     "C4:C8",
#     ColorScaleRule(start_type="min", start_color="00FF00", end_type="max", end_color="FF0000")
# )

# # Write Resource Breakdown Headers
# sheet1["A10"] = "Resource"
# sheet1["B10"] = "Cost (INR)"
# for col in range(1, 3):
#     cell = sheet1.cell(row=10, column=col)
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal="center")
#     cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# # Write Resource Breakdown Data
# for row, data in enumerate(resource_data, 11):
#     sheet1[f"A{row}"].value = data["resource"]
#     sheet1[f"B{row}"].value = data["cost"]
#     sheet1[f"A{row}"].alignment = Alignment(wrap_text=True)
#     sheet1[f"B{row}"].number_format = "₹#,##0.00"
#     for col in range(1, 3):
#         sheet1.cell(row=row, column=col).border = Border(
#             left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
#         )

# # Add Total
# sheet1["A15"] = "Total"
# sheet1["B15"] = sum(data["cost"] for data in resource_data)
# sheet1["A15"].font = Font(bold=True)
# sheet1["B15"].font = Font(bold=True)
# sheet1["B15"].number_format = "₹#,##0.00"
# sheet1["A15"].alignment = Alignment(horizontal="center")
# sheet1["B15"].alignment = Alignment(horizontal="center")

# # Adjust Column Widths
# sheet1.column_dimensions["A"].width = 50
# sheet1.column_dimensions["B"].width = 15
# sheet1.column_dimensions["C"].width = 15
# sheet1.column_dimensions["D"].width = 10

# # Add Pie Chart
# pie_chart = PieChart()
# data = Reference(sheet1, min_col=2, min_row=11, max_row=14)
# categories = Reference(sheet1, min_col=1, min_row=11, max_row=14)
# pie_chart.add_data(data)
# pie_chart.set_categories(categories)
# pie_chart.title = "POC-Immerz Resource Cost Distribution (INR)"
# pie_chart.dataLabels = pie_chart.dataLabels or openpyxl.chart.label.DataLabelList()
# pie_chart.dataLabels.showVal = True
# pie_chart.dataLabels.showPercent = True
# sheet1.add_chart(pie_chart, "E10")

# # --- Sheet2 ---
# sheet2 = wb.create_sheet("Sheet2")

# # Monthly Costs Data
# monthly_data2 = [
#     {"date": 45809, "usd": 36.4347816, "inr": 3109.981871, "currency": "INR"},
#     {"date": 45839, "usd": 85.87007505, "inr": 7359.387444, "currency": "INR"},
#     {"date": 45870, "usd": 84.76661605, "inr": 7359.437606, "currency": "INR"},
#     {"date": 45901, "usd": 19.87284024, "inr": 1742.599679, "currency": "INR"},
# ]

# # Resource Breakdown Data
# resource_data2 = [
#     {"resource": "migoto-saas-asp (microsoft.web/serverfarms)", "cost": 12749.47},
#     {"resource": "migoto-saas-sql / migoto-saasampsaasdb (microsoft.sql/servers/databases)", "cost": 3540.53},
#     {"resource": "migoto-saas-db-pe (microsoft.network/privateendpoints)", "cost": 1536.07},
#     {"resource": "migoto-saas-kv-pe (microsoft.network/privateendpoints)", "cost": 1536.04},
#     {"resource": "privatelink.database.windows.net (microsoft.network/privatednszones)", "cost": 103.23},
#     {"resource": "migoto-sql / migotoampsaasdb (microsoft.sql/servers/databases)", "cost": 1.61},
#     {"resource": "migoto-saas-portal (microsoft.web/sites)", "cost": 0.43},
#     {"resource": "migoto-saas-admin (microsoft.web/sites)", "cost": 0.42},
#     {"resource": "Others", "cost": 0.36},
# ]

# # Write Sheet Title
# sheet2["A1"] = "Sheet2 Azure Costs"
# sheet2["A1"].font = Font(bold=True, size=14)
# sheet2["A1"].fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")

# # Write Monthly Costs Headers
# for col, header in enumerate(monthly_headers, 1):
#     cell = sheet2.cell(row=3, column=col)
#     cell.value = header
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal="center")
#     cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# # Write Monthly Costs Data
# for row, data in enumerate(monthly_data2, 4):
#     sheet2[f"A{row}"] = base_date + timedelta(days=data["date"])
#     sheet2[f"B{row}"] = data["usd"]
#     sheet2[f"C{row}"] = data["inr"]
#     sheet2[f"D{row}"] = data["currency"]

# # Format Monthly Costs
# for row in range(4, 8):
#     sheet2[f"A{row}"].number_format = "yyyy-mm-dd"
#     sheet2[f"B{row}"].number_format = "$#,##0.00"
#     sheet2[f"C{row}"].number_format = "₹#,##0.00"
#     for col in range(1, 5):
#         sheet2.cell(row=row, column=col).alignment = Alignment(horizontal="center")
#         sheet2.cell(row=row, column=col).border = Border(
#             left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
#         )

# # Apply Conditional Formatting to Cost (INR)
# sheet2.conditional_formatting.add(
#     "C4:C7",
#     ColorScaleRule(start_type="min", start_color="00FF00", end_type="max", end_color="FF0000")
# )

# # Write Resource Breakdown Headers
# sheet2["A9"] = "Resource"
# sheet2["B9"] = "Cost (INR)"
# for col in range(1, 3):
#     cell = sheet2.cell(row=9, column=col)
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal="center")
#     cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

# # Write Resource Breakdown Data
# for row, data in enumerate(resource_data2, 10):
#     sheet2[f"A{row}"].value = data["resource"]
#     sheet2[f"B{row}"].value = data["cost"]
#     sheet2[f"A{row}"].alignment = Alignment(wrap_text=True)
#     sheet2[f"B{row}"].number_format = "₹#,##0.00"
#     for col in range(1, 3):
#         sheet2.cell(row=row, column=col).border = Border(
#             left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
#         )

# # Add Total
# sheet2["A18"] = "Total"
# sheet2["B18"] = sum(data["cost"] for data in resource_data2)
# sheet2["A18"].font = Font(bold=True)
# sheet2["B18"].font = Font(bold=True)
# sheet2["B18"].number_format = "₹#,##0.00"
# sheet2["A18"].alignment = Alignment(horizontal="center")
# sheet2["B18"].alignment = Alignment(horizontal="center")

# # Adjust Column Widths
# sheet2.column_dimensions["A"].width = 50
# sheet2.column_dimensions["B"].width = 15
# sheet2.column_dimensions["C"].width = 15
# sheet2.column_dimensions["D"].width = 10

# # Add Bar Chart
# bar_chart = BarChart()
# data = Reference(sheet2, min_col=2, min_row=10, max_row=18)
# categories = Reference(sheet2, min_col=1, min_row=10, max_row=18)
# bar_chart.add_data(data)
# bar_chart.set_categories(categories)
# bar_chart.title = "Sheet2 Resource Cost Breakdown (INR)"
# bar_chart.dataLabels = bar_chart.dataLabels or openpyxl.chart.label.DataLabelList()
# bar_chart.dataLabels.showVal = True
# sheet2.add_chart(bar_chart, "E9")

# # Save the Workbook
# wb.save("Azure_Costing_Pretty.xlsx")
# print("Excel file 'Azure_Costing_Pretty.xlsx' has been created successfully!")
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import PieChart, Reference
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()

# Remove the default sheet
wb.remove(wb.active)

# --- migoto-devp Sheet ---
sheet = wb.create_sheet("migoto-devp")

# Monthly Costs Data
monthly_data = [
    {"date": datetime(2025, 7, 1), "usd": 2.480223847, "inr": 212.5644846, "currency": "INR"},
    {"date": datetime(2025, 8, 1), "usd": 17.86965996, "inr": 1551.443878, "currency": "INR"},
]

# Resource Breakdown Data
resource_data = [
    {"resource": "microsoft.compute/virtualmachines", "cost": 1447.78},
    {"resource": "microsoft.compute/disks", "cost": 316.23},
]

# Write Sheet Title
sheet["A1"] = "migoto-devp Azure Costs"
sheet["A1"].font = Font(bold=True, size=14)
sheet["A1"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Write Monthly Costs Headers
monthly_headers = ["Usage Date (Billing Month)", "Cost (USD)", "Cost (INR)", "Currency"]
for col, header in enumerate(monthly_headers, 1):
    cell = sheet.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Write Monthly Costs Data
for row, data in enumerate(monthly_data, 4):
    sheet[f"A{row}"] = data["date"]
    sheet[f"B{row}"] = data["usd"]
    sheet[f"C{row}"] = data["inr"]
    sheet[f"D{row}"] = data["currency"]

# Format Monthly Costs
for row in range(4, 6):
    sheet[f"A{row}"].number_format = "dd-mm-yyyy"
    sheet[f"B{row}"].number_format = "$#,##0.00"
    sheet[f"C{row}"].number_format = "₹#,##0.00"
    for col in range(1, 5):
        sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        sheet.cell(row=row, column=col).border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

# Apply Conditional Formatting to Cost (INR)
sheet.conditional_formatting.add(
    "C4:C5",
    ColorScaleRule(start_type="min", start_color="00FF00", end_type="max", end_color="FF0000")
)

# Write Resource Breakdown Headers
sheet["A7"] = "Resource"
sheet["B7"] = "Cost (INR)"
for col in range(1, 3):
    cell = sheet.cell(row=7, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Write Resource Breakdown Data
for row, data in enumerate(resource_data, 8):
    sheet[f"A{row}"].value = data["resource"]
    sheet[f"B{row}"].value = data["cost"]
    sheet[f"A{row}"].alignment = Alignment(wrap_text=True)
    sheet[f"B{row}"].number_format = "₹#,##0.00"
    for col in range(1, 3):
        sheet.cell(row=row, column=col).border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

# Add Total
sheet["A10"] = "Total"
sheet["B10"] = sum(data["cost"] for data in resource_data)
sheet["A10"].font = Font(bold=True)
sheet["B10"].font = Font(bold=True)
sheet["B10"].number_format = "₹#,##0.00"
sheet["A10"].alignment = Alignment(horizontal="center")
sheet["B10"].alignment = Alignment(horizontal="center")

# Adjust Column Widths
sheet.column_dimensions["A"].width = 40
sheet.column_dimensions["B"].width = 15
sheet.column_dimensions["C"].width = 15
sheet.column_dimensions["D"].width = 10

# Add Pie Chart
pie_chart = PieChart()
data = Reference(sheet, min_col=2, min_row=8, max_row=9)
categories = Reference(sheet, min_col=1, min_row=8, max_row=9)
pie_chart.add_data(data)
pie_chart.set_categories(categories)
pie_chart.title = "migoto-devp Resource Cost Distribution (INR)"
pie_chart.dataLabels = pie_chart.dataLabels or openpyxl.chart.label.DataLabelList()
pie_chart.dataLabels.showVal = True
pie_chart.dataLabels.showPercent = True
sheet.add_chart(pie_chart, "E7")

# Save the Workbook
wb.save("Migoto_Devp_Pretty.xlsx")
print("Excel file 'Migoto_Devp_Pretty.xlsx' has been created successfully!")